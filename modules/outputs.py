"""
Output generation module — charts, CSV builders, and summary formatters.
"""

import calendar
import math
from html import escape as _esc

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from io import StringIO, BytesIO
from .billing import BillingResult, simulate_year_under_billing_option


_NEM12_REGIMES = ("NEM-1", "NEM-2", "NEM-A (NEM-1)", "NEM-A (NEM-2)")
_NEM3_REGIMES = ("NEM-3", "NEM-3 / NVBT", "NEM-A (NEM-3)", "NEM-A (NEM-3 / NVBT)")


# 38DN Excel number formats — four-section accounting layout per
# conventions.md: positive normal, negative parentheses, zero en-dash,
# text passthrough. Imported by other export modules so the entire
# pipeline renders zeros consistently.
EXCEL_FMT_KWH = '_(#,##0_);(#,##0);_("–"_);_(@_)'
EXCEL_FMT_DOLLAR = '_($#,##0_);($#,##0);_($"–"_);_(@_)'
EXCEL_FMT_DOLLAR_ACCT = '_($#,##0_);[Red]($#,##0);_($"–"_);_(@_)'
EXCEL_FMT_DOLLAR_K = '_($#,##0"K"_);($#,##0"K");_($"–"_);_(@_)'
EXCEL_FMT_RATE = '_($0.00000_);($0.00000);_($"–"_);_(@_)'
EXCEL_FMT_PCT = '_(0.0%_);(0.0%);_("–"_);_(@_)'


def _supports_nsc(regime: str | None) -> bool:
    """All CA NEM regimes carry an annual NSC true-up — NEM-1/2 re-prices
    surplus from retail TOU to NSC, NEM-3/NBT re-prices from avg ACC to NSC.
    """
    if not regime:
        return False
    return regime in _NEM12_REGIMES or regime in _NEM3_REGIMES


def _compute_export_cagr(multiyear: dict[int, "pd.Series"], n_trailing: int = 10) -> float:
    """Compute the CAGR of average export rates over the last n_trailing years of a CSV.

    Returns the annual growth rate (e.g. 0.02 for 2%/yr). Returns 0.0 if
    fewer than 2 years of data are available.
    """
    keys = sorted(multiyear.keys())
    if len(keys) < 2:
        return 0.0
    # Use the last n_trailing years (or all years if fewer available)
    tail_keys = keys[-n_trailing:] if len(keys) >= n_trailing else keys
    if len(tail_keys) < 2:
        return 0.0
    first_yr = tail_keys[0]
    last_yr = tail_keys[-1]
    avg_first = float(multiyear[first_yr].mean())
    avg_last = float(multiyear[last_yr].mean())
    if avg_first <= 0:
        return 0.0
    span = last_yr - first_yr
    if span <= 0:
        return 0.0
    return (avg_last / avg_first) ** (1.0 / span) - 1.0


MONTH_NAMES = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]


# ---------------------------------------------------------------------------
# Negative-value formatting helpers (accounting style)
# ---------------------------------------------------------------------------
def fmt_num(x) -> str:
    """Format a number as XXX,XXX with parentheses for negatives."""
    if not isinstance(x, (int, float)):
        return str(x)
    if x < 0:
        return f"({abs(x):,.0f})"
    return f"{x:,.0f}"


def fmt_dollar(x) -> str:
    """Format a number as $XXX,XXX with parentheses for negatives.

    NaN renders as an em-dash to flag "not applicable" cells (e.g., NSC Adj
    in NEM-3 years where no annual true-up exists). $0 still renders as $0.
    """
    if not isinstance(x, (int, float)):
        return str(x)
    if math.isnan(x):
        return "—"
    if x < 0:
        return f"$({abs(x):,.0f})"
    return f"${x:,.0f}"


def fmt_rate(x) -> str:
    """Format a number as $0.XXXXX rate with parentheses for negatives."""
    if not isinstance(x, (int, float)):
        return str(x)
    if x < 0:
        return f"$({abs(x):.5f})"
    return f"${x:.5f}"


def style_negative_red(styler):
    """Apply red styling to any cell whose text contains '(' (accounting negative)."""
    def _color(val):
        if isinstance(val, str) and "(" in val:
            # WCAG-friendly: darker red text, paler background (was #cc0000/#ffe0e0).
            return "color: #a8141a; background-color: #fff5f5"
        return ""
    return styler.map(_color)


def render_styled_table(
    df: pd.DataFrame,
    bold_last_row: bool = False,
    bold_cols: list[str] | None = None,
    highlight_cols: list[str] | None = None,
) -> str:
    """Render a DataFrame as an HTML table using the 38DN institutional
    table style (see ``assets/theme.css::.tbl-38dn``).

    Classes are applied on cells; the stylesheet handles colors, typography,
    zebra stripes, hover, and negatives. Output is class-based — re-skinning
    only requires editing ``assets/theme.css``.

    Args:
        df: pre-formatted DataFrame (every cell is a display-ready string).
        bold_last_row: marks the last row as a TOTAL (navy top border + weight).
        bold_cols: columns whose header + cells render semibold.
        highlight_cols: columns rendered with a subtle info-blue background
            (reserved for "pay attention here" totals like Cumulative Savings).
    """
    bold_set = set(bold_cols) if bold_cols else set()
    highlight_set = set(highlight_cols) if highlight_cols else set()
    col_list = list(df.columns)

    html = ['<div class="tbl-38dn">', "<table>", "<thead><tr>"]
    for col in col_list:
        extra = ' class="cell-bold"' if col in bold_set else ""
        html.append(f"<th{extra}>{_esc(str(col))}</th>")
    html.append("</tr></thead><tbody>")

    for i, (_, row) in enumerate(df.iterrows()):
        is_total = bold_last_row and i == len(df) - 1
        tr_cls = ' class="total-row"' if is_total else ""
        html.append(f"<tr{tr_cls}>")
        for j, val in enumerate(row):
            s = str(val)
            # Pre-rendered SVG (sparklines) passes through unescaped. Every
            # other cell is HTML-escaped to prevent injection.
            is_raw_svg = s.lstrip().startswith("<svg")
            classes: list[str] = []
            if col_list[j] in bold_set or is_total:
                classes.append("cell-bold")
            if col_list[j] in highlight_set:
                classes.append("cell-highlight")
            elif "(" in s and not is_raw_svg:
                # Accounting negative: `$(1,234)` format rendered red.
                classes.append("cell-negative")
            cls_attr = f' class="{" ".join(classes)}"' if classes else ""
            cell_html = s if is_raw_svg else _esc(s)
            html.append(f"<td{cls_attr}>{cell_html}</td>")
        html.append("</tr>")

    html.append("</tbody></table></div>")
    return "\n".join(html)


def _negate_outflow_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Negate cost outflow columns (accounting style) for display / CSV.

    Cost components only — kWh columns (Export (kWh), Import (kWh), etc.)
    are quantities, not dollar flows, and stay positive. The on-screen
    Annual Projection table uses the same rule; this helper exists so the
    downloaded CSV matches what the user sees without recomputing the
    negation inline at the download call site.
    """
    out = df.copy()
    for col in ["Bill w/o Solar ($)", "Energy ($)", "Demand ($)",
                 "Fixed ($)", "NBC ($)", "NSC Adj ($)", "Bill w/ Solar ($)"]:
        if col in out.columns:
            out[col] = out[col] * -1
    return out


def build_monthly_summary_display(
    result: BillingResult,
    result_pv_only: BillingResult | None = None,
    existing_solar_offset_kwh: list[float] | None = None,
) -> pd.DataFrame:
    """
    Format the monthly summary for display in Streamlit.
    Adds month names and formats currency columns.

    If result_pv_only is provided, the table includes both
    "Demand kW (PV)" (from result_pv_only) and "Demand kW (PV+BESS)" (from result).
    Otherwise, only "Demand kW (PV)" is shown (from result itself).
    """
    df = result.monthly_summary.copy()
    df["month_name"] = [MONTH_NAMES[int(m) - 1] for m in df["month"]]

    # Build display columns and rename map
    display_cols = [
        "month_name", "solar_kwh", "import_kwh", "export_kwh",
        "export_peak_kwh", "export_offpeak_kwh",
    ]
    rename_map = {
        "month_name": "Month",
        "solar_kwh": "Solar (kWh)",
        "import_kwh": "Import (kWh)",
        "export_kwh": "Export (kWh)",
        "export_peak_kwh": "Export Peak (kWh)",
        "export_offpeak_kwh": "Export Off-Peak (kWh)",
    }

    # Conditionally add degraded system load offset column
    if existing_solar_offset_kwh is not None and len(existing_solar_offset_kwh) == 12:
        df["degraded_offset_kwh"] = existing_solar_offset_kwh
        display_cols.insert(1, "degraded_offset_kwh")
        rename_map["degraded_offset_kwh"] = "Degraded System Load Offset (kWh)"

    if result_pv_only is not None:
        # BESS mode: show both PV-only and PV+BESS demand
        df["demand_kw_pv"] = result_pv_only.monthly_summary["peak_demand_kw"]
        df["demand_kw_bess"] = df["peak_demand_kw"]
        display_cols += ["demand_kw_pv", "demand_kw_bess"]
        rename_map["demand_kw_pv"] = "Demand kW (PV)"
        rename_map["demand_kw_bess"] = "Demand kW (PV+BESS)"
    else:
        # PV-only mode
        display_cols.append("peak_demand_kw")
        rename_map["peak_demand_kw"] = "Demand kW (PV)"

    display_cols += [
        "energy_cost", "total_demand_charge",
        "fixed_charge",
    ]

    # NBC column (NEM-2 only — include when any month has nbc_charge > 0)
    _has_nbc = "nbc_charge" in result.monthly_summary.columns and result.monthly_summary["nbc_charge"].sum() > 0
    if _has_nbc:
        display_cols.append("nbc_charge")

    display_cols += ["export_credit", "net_bill"]

    # Rate shift savings column (when old-rate baseline is available)
    if result.old_rate_monthly_baselines is not None and result.monthly_baseline_details is not None:
        _old_baselines = result.old_rate_monthly_baselines
        _new_baselines = [d["total"] for d in result.monthly_baseline_details]
        df["rate_shift_savings"] = [
            round(_old_baselines[i] - _new_baselines[i], 2)
            for i in range(12)
        ]
        display_cols.append("rate_shift_savings")

    # Sub-component export kWh columns indented under the total Export (kWh) column.
    rename_map.update({
        "energy_cost": "Energy ($)",
        "total_demand_charge": "Demand ($)",
        "fixed_charge": "Fixed ($)",
        "nbc_charge": "NBC ($)",
        "export_credit": "Export Credit ($)",
        "net_bill": "Net Bill ($)",
        "rate_shift_savings": "Rate Shift Savings ($)",
        "export_peak_kwh": "↳ Peak (kWh)",
        "export_offpeak_kwh": "↳ Off-Peak (kWh)",
    })

    # Only include columns that exist
    display_cols = [c for c in display_cols if c in df.columns]
    df = df[display_cols].rename(columns=rename_map)

    # Negate cost-outflow columns so the bill rows render as accounting negatives
    # (Energy/Demand/Fixed/NBC/Net Bill as red parentheses, Export Credit as
    # positive). Export Credit and Rate Shift Savings stay positive.
    COST_OUTFLOWS = ("Energy ($)", "Demand ($)", "Fixed ($)", "NBC ($)", "Net Bill ($)")
    for col in COST_OUTFLOWS:
        if col in df.columns:
            df[col] = df[col] * -1

    # Format kWh columns
    kwh_cols = [c for c in df.columns if "(kWh)" in c]
    for col in kwh_cols:
        df[col] = df[col].apply(fmt_num)

    # Format kW columns
    kw_cols = [c for c in df.columns if "kW" in c and "(kWh)" not in c]
    for col in kw_cols:
        df[col] = df[col].apply(fmt_num)

    # Format $ columns
    dollar_cols = [c for c in df.columns if "($)" in c]
    for col in dollar_cols:
        df[col] = df[col].apply(fmt_dollar)

    return df


def build_savings_summary(result: BillingResult, system_cost: float = 0.0) -> dict:
    """Build a savings summary dictionary."""
    simple_payback = None
    if result.annual_savings > 0 and system_cost > 0:
        simple_payback = system_cost / result.annual_savings

    summary = {
        "annual_load_kwh": round(result.annual_load_kwh, 0),
        "annual_solar_kwh": round(result.annual_solar_kwh, 0),
        "solar_offset_pct": round(
            result.annual_solar_kwh / result.annual_load_kwh * 100
            if result.annual_load_kwh > 0 else 0, 1
        ),
        "annual_import_kwh": round(result.annual_import_kwh, 0),
        "annual_export_kwh": round(result.annual_export_kwh, 0),
        "annual_bill_without_solar": round(result.annual_bill_without_solar, 2),
        "annual_bill_with_solar": round(result.annual_bill_with_solar, 2),
        "annual_savings": round(result.annual_savings, 2),
        "savings_pct": round(result.savings_pct, 1),
        "system_cost": round(system_cost, 2),
        "simple_payback_years": round(simple_payback, 1) if simple_payback else None,
    }

    # Rate shift analysis fields
    if result.rate_shift_annual_savings is not None:
        summary["rate_shift_annual_savings"] = round(result.rate_shift_annual_savings, 2)
        summary["total_annual_savings"] = round(
            result.annual_savings + result.rate_shift_annual_savings, 2
        )

    return summary


def _compute_tou_netted_monthly(hourly_detail: pd.DataFrame) -> tuple[
    float, float, dict[int, float], dict[int, float]
]:
    """Compute TOU-netted energy cost and export credit from hourly data.

    .. deprecated::
        Annual totals are now pre-computed in BillingResult.tou_annual_energy
        and BillingResult.tou_annual_credit.  This function is retained only
        for callers that need the per-month breakdown dicts.

    Returns:
        (annual_tou_energy, annual_tou_credit,
         per_month_tou_energy, per_month_tou_credit)

    Under TOU netting (NEM-1/2), each month×period's import and export are
    netted at the TOU retail rate. Months with net-positive charges produce
    energy cost; months with net-negative charges produce export credit.
    """
    annual_energy = 0.0
    annual_credit = 0.0
    month_energy: dict[int, float] = {}
    month_credit: dict[int, float] = {}

    has_rate = "energy_rate" in hourly_detail.columns
    for month in range(1, 13):
        mm = hourly_detail.index.month == month
        if not has_rate:
            # ECC fallback
            ec = float(hourly_detail.loc[mm, "export_credit"].sum()) if "export_credit" in hourly_detail.columns else 0.0
            month_energy[month] = 0.0
            month_credit[month] = ec
            annual_credit += ec
            continue

        monthly_charge = 0.0
        for pidx in hourly_detail.loc[mm, "energy_period"].unique():
            pm = mm & (hourly_detail["energy_period"] == pidx)
            rate = hourly_detail.loc[pm, "energy_rate"].iloc[0]
            net = hourly_detail.loc[pm, "import_kwh"].sum() - hourly_detail.loc[pm, "export_kwh"].sum()
            monthly_charge += net * rate

        if monthly_charge >= 0:
            month_energy[month] = monthly_charge
            month_credit[month] = 0.0
            annual_energy += monthly_charge
        else:
            month_energy[month] = 0.0
            month_credit[month] = abs(monthly_charge)
            annual_credit += abs(monthly_charge)

    return annual_energy, annual_credit, month_energy, month_credit


def _resolve_regime_context(
    yr: int,
    nem_regime_1: str,
    nem_regime_2: str | None,
    num_years_1: int | None,
    export_rates_multiyear: dict | None,
    export_rates_multiyear_2: dict | None,
    multiyear_start: int,
    multiyear_max: int,
    multiyear_start_2: int,
    multiyear_max_2: int,
    export_cagr: float,
    export_cagr_2: float,
) -> tuple[str, dict | None, int, int, float]:
    """Determine the active NEM regime and export rate context for a given year.

    Returns:
        (active_regime, active_multiyear, active_my_start, active_my_max, active_cagr)
    """
    if nem_regime_2 and num_years_1 and yr > num_years_1:
        return (nem_regime_2, export_rates_multiyear_2,
                multiyear_start_2, multiyear_max_2, export_cagr_2)
    return (nem_regime_1, export_rates_multiyear,
            multiyear_start, multiyear_max, export_cagr)


def _compute_year_row(
    yr: int,
    *,
    year1_load_kwh: float,
    year1_solar_kwh: float,
    year1_import_kwh: float,
    year1_export_kwh: float,
    year1_export_peak_kwh: float,
    year1_export_offpeak_kwh: float,
    year1_demand: float,
    year1_fixed: float,
    year1_export: float,
    year1_nbc: float,
    year1_demand_kw_pv: float,
    year1_demand_kw_bess: float,
    year1_bill_no_solar: float,
    year1_baseline_energy: float,
    year1_baseline_demand: float,
    year1_baseline_fixed: float,
    tou_year1_energy: float,
    year1_tou_credit: float,
    gen_raw_energy: float,
    agg_raw_energy: float,
    blended_import_rate: float,
    rate_mult: float,
    load_mult: float,
    degrad_rate: float,
    compound_escalation: bool,
    active_regime: str,
    active_multiyear: dict | None,
    active_my_start: int,
    active_my_max: int,
    active_cagr: float,
    cod_year: int | None,
    nem_regime_1: str,
    nem_regime_2: str | None,
    num_years_1: int | None,
    nbc_rate_2: float,
    nsc_rate_2: float,
    result_pv_only,
    result_hourly_detail,
    result_annual_bill_with_solar: float,
    existing_solar_offset_kwh: float,
    year1_nsc_adj: float = 0.0,
    monthly_summary_y1: pd.DataFrame | None = None,
    tou_monthly_energy_y1: dict | None = None,
    tou_monthly_credit_y1: dict | None = None,
    billing_option: str = "MBO",
    min_monthly_charge: float = 0.0,
) -> dict:
    """Compute a single year's projection row for build_annual_projection."""
    if compound_escalation:
        rate_factor = (1.0 + rate_mult) ** (yr - 1)
        load_factor = (1.0 + load_mult) ** (yr - 1)
    else:
        rate_factor = 1.0 + rate_mult * (yr - 1)
        load_factor = 1.0 + load_mult * (yr - 1)

    solar_factor = (1.0 - degrad_rate) ** (yr - 1)

    yr_load_kwh = year1_load_kwh * load_factor
    yr_solar_kwh = year1_solar_kwh * solar_factor

    year1_self_consumption = year1_solar_kwh - year1_export_kwh
    if year1_solar_kwh > 0:
        self_consumption_frac = year1_self_consumption / year1_solar_kwh
        export_frac = year1_export_kwh / year1_solar_kwh
    else:
        self_consumption_frac = 1.0
        export_frac = 0.0

    lost_solar = year1_solar_kwh * (1.0 - solar_factor)
    extra_load = year1_load_kwh * (load_factor - 1)
    lost_export_from_degrad = lost_solar * export_frac
    lost_self_from_degrad = lost_solar * self_consumption_frac
    lost_export_from_load = min(extra_load, max(0, year1_export_kwh - lost_export_from_degrad))

    yr_export_kwh = max(0, year1_export_kwh - lost_export_from_degrad - lost_export_from_load)
    yr_import_kwh = year1_import_kwh + lost_self_from_degrad + max(0, extra_load - lost_export_from_load)

    export_volume_ratio = yr_export_kwh / year1_export_kwh if year1_export_kwh > 0 else 0.0
    yr_export_peak_kwh = year1_export_peak_kwh * export_volume_ratio
    yr_export_offpeak_kwh = year1_export_offpeak_kwh * export_volume_ratio

    yr_demand_kw_pv = year1_demand_kw_pv * load_factor
    yr_demand_kw_bess = year1_demand_kw_bess * load_factor

    yr_demand = year1_demand * load_factor * rate_factor
    yr_fixed = year1_fixed * rate_factor
    volume_ratio = yr_export_kwh / year1_export_kwh if year1_export_kwh > 0 else 0.0
    import_ratio = yr_import_kwh / year1_import_kwh if year1_import_kwh > 0 else load_factor

    _is_tou_netted = active_regime in ("NEM-1", "NEM-2") or active_regime.startswith("NEM-A")
    if _is_tou_netted:
        degradation_energy_adj = lost_self_from_degrad * blended_import_rate * rate_factor
        yr_energy = tou_year1_energy * load_factor * rate_factor + degradation_energy_adj
    else:
        yr_energy = (
            gen_raw_energy * import_ratio * rate_factor
            + agg_raw_energy * load_factor * rate_factor
        )

    if _is_tou_netted:
        yr_export_credit = year1_tou_credit * rate_factor * volume_ratio
    elif active_my_max > 0 and active_multiyear:
        calendar_year = (cod_year if cod_year is not None else active_my_start) + (yr - 1)
        rate_year = max(active_my_start, min(calendar_year, active_my_max))
        yr_export_rates = active_multiyear[rate_year].values
        hourly_export = result_hourly_detail["export_kwh"].values
        base_credit = float(np.sum(hourly_export * yr_export_rates))
        if calendar_year > active_my_max and active_cagr != 0.0:
            overshoot = calendar_year - active_my_max
            base_credit *= (1.0 + active_cagr) ** overshoot
        yr_export_credit = base_credit * volume_ratio
    elif nem_regime_2 and num_years_1 and yr > num_years_1:
        yr_export_credit = 0.0
    else:
        if _is_tou_netted:
            yr_export_credit = year1_export * rate_factor * volume_ratio
        else:
            yr_export_credit = year1_export * volume_ratio

    if active_regime == "NEM-2" or active_regime == "NEM-A (NEM-2)":
        if nem_regime_2 and num_years_1 and yr > num_years_1 and nbc_rate_2 > 0:
            yr_nbc = nbc_rate_2 * yr_import_kwh * rate_factor
        else:
            yr_nbc = year1_nbc * rate_factor if year1_nbc > 0 else 0.0
    else:
        yr_nbc = 0.0

    yr_nsc = 0.0
    if nem_regime_2 and num_years_1 and yr > num_years_1 and nsc_rate_2 > 0:
        if active_regime in ("NEM-1", "NEM-2", "NEM-A (NEM-1)", "NEM-A (NEM-2)"):
            yr_nsc = nsc_rate_2 * yr_export_kwh * rate_factor

    # End-of-year NSC clawback: applies under NEM-1/2 (retail TOU → NSC) and
    # under NEM-3/NBT (avg ACC → NSC). The Y1 delta sits in
    # result.annual_nsc_adjustment; year-N scaling differs by regime:
    #   NEM-1/2: rate × volume — retail TOU credit and NSC payment move with
    #            the utility rate escalator and export volume.
    #   NEM-3:   volume only — ACC export rates follow the loaded ACC schedule
    #            (already reflected in yr_export_credit), not the retail
    #            escalator. NSC ($/kWh) is wholesale and roughly stable in
    #            real terms, so volume_ratio is the dominant scaler.
    yr_nsc_clawback = 0.0
    if year1_nsc_adj > 0 and _supports_nsc(active_regime):
        if active_regime in _NEM12_REGIMES:
            yr_nsc_clawback = year1_nsc_adj * rate_factor * volume_ratio
        else:
            yr_nsc_clawback = year1_nsc_adj * volume_ratio

    # For Y>1 in NEM-1/2 (single-meter or NEM-A aggregate), re-run MBO/ABO
    # month-by-month so credit banking + min_monthly_charge floors land where
    # they should. Without this, over-sized PV under MBO sees gross TOU credit
    # subtract from gross charges at full magnitude, which understates the bill
    # for any year where monthly bills would have floored at $0. For NEM-A,
    # the helper applies MBO at the aggregate level using the aggregate TOU
    # split derived in _build_aggregate_result; aggregate Y1 net_bill itself
    # doesn't apply aggregate-level MBO, so Y1 and Y>1 may diverge for
    # portfolios where aggregate monthly net would have floored.
    _can_simulate_year = (
        yr > 1
        and active_regime in ("NEM-1", "NEM-2", "NEM-A (NEM-1)", "NEM-A (NEM-2)")
        and tou_monthly_energy_y1 is not None
        and tou_monthly_credit_y1 is not None
        and monthly_summary_y1 is not None
    )
    if _can_simulate_year:
        monthly_components = []
        total_load_y1 = float(monthly_summary_y1["load_kwh"].sum()) or 1.0
        _has_nbc_in_regime = active_regime in ("NEM-2", "NEM-A (NEM-2)")
        for mi in range(12):
            ms_row = monthly_summary_y1.iloc[mi]
            m_energy = tou_monthly_energy_y1.get(mi + 1, 0.0) * load_factor * rate_factor
            m_credit = tou_monthly_credit_y1.get(mi + 1, 0.0) * rate_factor * volume_ratio
            m_demand = float(ms_row["total_demand_charge"]) * load_factor * rate_factor
            m_fixed = float(ms_row["fixed_charge"]) * rate_factor
            m_nbc = float(ms_row.get("nbc_charge", 0.0)) * rate_factor if _has_nbc_in_regime else 0.0
            # NSC clawback + regime-2 NSC charge both land on month 12 of the bill
            m_nsc = (yr_nsc_clawback + yr_nsc) if mi == 11 else 0.0
            monthly_components.append({
                "energy": m_energy,
                "export_credit": m_credit,
                "demand": m_demand,
                "fixed": m_fixed,
                "nbc": m_nbc,
                "nsc_adj": m_nsc,
            })
        # Distribute the annual degradation-energy adjustment proportional to monthly load
        if degradation_energy_adj > 0:
            for mi in range(12):
                mo_load_share = float(monthly_summary_y1.iloc[mi]["load_kwh"]) / total_load_y1
                monthly_components[mi]["energy"] += degradation_energy_adj * mo_load_share
        net_bills = simulate_year_under_billing_option(
            monthly_components, active_regime, billing_option, min_monthly_charge,
        )
        yr_bill_solar_raw = sum(net_bills)
    else:
        yr_bill_solar_raw = yr_energy + yr_demand + yr_fixed + yr_nbc + yr_nsc + yr_nsc_clawback - yr_export_credit

    if yr == 1:
        yr_bill_solar = result_annual_bill_with_solar
    else:
        yr_bill_solar = yr_bill_solar_raw

    yr_baseline_energy = year1_baseline_energy * load_factor * rate_factor
    yr_baseline_demand = year1_baseline_demand * load_factor * rate_factor
    yr_baseline_fixed = year1_baseline_fixed * rate_factor
    yr_bill_no_solar_raw = yr_baseline_energy + yr_baseline_demand + yr_baseline_fixed
    if yr == 1:
        yr_bill_no_solar = year1_bill_no_solar
    else:
        yr_bill_no_solar = yr_bill_no_solar_raw

    yr_savings = yr_bill_no_solar - yr_bill_solar

    row = {
        "Year": yr,
    }
    if cod_year is not None:
        row["Calendar Year"] = cod_year + yr - 1
    yr_self_consumed = yr_solar_kwh - max(yr_export_kwh, 0)
    if existing_solar_offset_kwh > 0:
        row["Degraded System Load Offset (kWh)"] = round(existing_solar_offset_kwh * load_factor)
    row.update({
        "Customer Load (kWh)": round(yr_load_kwh),
        "Solar (kWh)": round(yr_solar_kwh),
        "Solar Offset (kWh)": round(yr_self_consumed),
        "Import (kWh)": round(yr_import_kwh),
        "Export (kWh)": round(max(yr_export_kwh, 0)),
        "Export Peak (kWh)": round(yr_export_peak_kwh),
        "Export Off-Peak (kWh)": round(yr_export_offpeak_kwh),
        "Demand kW (PV)": round(yr_demand_kw_pv),
    })
    if result_pv_only is not None:
        row["Demand kW (PV+BESS)"] = round(yr_demand_kw_bess)
    row.update({
        "Bill w/o Solar ($)": round(yr_bill_no_solar),
        "Energy ($)": round(yr_energy),
        "Demand ($)": round(yr_demand),
        "Fixed ($)": round(yr_fixed),
    })
    _any_nem2 = any(r in ("NEM-2", "NEM-A (NEM-2)") for r in (nem_regime_1, nem_regime_2) if r)
    if _any_nem2 or year1_nbc > 0:
        row["NBC ($)"] = round(yr_nbc)
    # Populate NSC Adj column when ANY year in the projection carries NSC
    # (NEM-1/2 retail-to-NSC repricing OR NEM-3/NBT ACC-to-NSC repricing).
    # Years where the active regime doesn't support NSC get NaN (renders as
    # em-dash via fmt_dollar); applicable years get the value (or 0.0 when
    # no surplus). Conditional column population would create NaN holes via
    # pd.DataFrame and display "$nan".
    _any_nsc_regime = any(_supports_nsc(r) for r in (nem_regime_1, nem_regime_2))
    if _any_nsc_regime:
        yr_nsc_total = yr_nsc + yr_nsc_clawback
        row["NSC Adj ($)"] = round(yr_nsc_total) if _supports_nsc(active_regime) else float("nan")
    row.update({
        "Export Credit ($)": round(yr_export_credit),
        "Bill w/ Solar ($)": round(yr_bill_solar),
        "Annual Savings ($)": round(yr_savings),
    })

    return row


def _apply_rate_shift_columns(
    rows: list[dict],
    rate_shift_old_baseline: float,
    rate_mult: float,
    load_mult: float,
    compound_escalation: bool,
) -> None:
    """Add rate shift analysis columns to projection rows in-place."""
    cumulative_total_savings = 0.0
    for row in rows:
        yr = row["Year"]
        if compound_escalation:
            rate_factor = (1.0 + rate_mult) ** (yr - 1)
            load_factor = (1.0 + load_mult) ** (yr - 1)
        else:
            rate_factor = 1.0 + rate_mult * (yr - 1)
            load_factor = 1.0 + load_mult * (yr - 1)
        yr_old_rate_baseline = rate_shift_old_baseline * load_factor * rate_factor
        yr_rate_shift_savings = yr_old_rate_baseline - row["Bill w/o Solar ($)"]
        yr_total_savings = row["Annual Savings ($)"] + yr_rate_shift_savings
        cumulative_total_savings += yr_total_savings
        row["Old Rate Baseline ($)"] = round(yr_old_rate_baseline)
        row["Rate Shift Savings ($)"] = round(yr_rate_shift_savings)
        row["Total Savings ($)"] = round(yr_total_savings)
        row["Cumulative Total Savings ($)"] = round(cumulative_total_savings)


def build_annual_projection(
    result: BillingResult,
    system_cost: float,
    rate_escalator_pct: float,
    load_escalator_pct: float,
    years: int = 10,
    export_rates_multiyear: dict[int, "pd.Series"] | None = None,
    result_pv_only: BillingResult | None = None,
    nem_regime_1: str = "NEM-3 / NVBT",
    nem_regime_2: str | None = None,
    num_years_1: int | None = None,
    export_rates_multiyear_2: dict[int, "pd.Series"] | None = None,
    cod_year: int | None = None,
    degradation_pct: float = 0.0,
    nbc_rate_2: float = 0.0,
    nsc_rate_2: float = 0.0,
    compound_escalation: bool = True,
    rate_shift_old_baseline: float | None = None,
    existing_solar_offset_kwh: float = 0.0,
) -> pd.DataFrame:
    """
    Build a multi-year annual projection table.

    Escalators:
      - rate_escalator_pct: applied to TOU energy rates each year
      - load_escalator_pct: applied to load profile each year
        (increases energy consumption AND peak demand)
      - degradation_pct: annual solar production decline (e.g. 0.5 for 0.5%/yr)
      - compound_escalation: if True, use compound formula (1+r)^n;
        if False, use linear formula 1 + r*n

    Args:
        result: Year-1 BillingResult
        system_cost: Total installed cost ($)
        rate_escalator_pct: Annual TOU rate escalation (e.g., 3.0 for 3%)
        load_escalator_pct: Annual load/demand growth (e.g., 2.0 for 2%)
        years: Number of years to project
        result_pv_only: If provided, PV-only result for separate demand column
        nem_regime_1: NEM regime for the first period (default NEM-3/NVBT)
        nem_regime_2: NEM regime for the second period (None if no switch)
        num_years_1: Number of years under regime 1 (None if no switch)
        export_rates_multiyear_2: Multi-year export rates for regime 2
        compound_escalation: Use compound (True) or linear (False) escalation
    """
    year1_energy = float(result.monthly_summary["energy_cost"].sum())
    year1_demand = float(result.monthly_summary["total_demand_charge"].sum())
    year1_fixed = float(result.monthly_summary["fixed_charge"].sum())
    year1_export = float(result.monthly_summary["export_credit"].sum())
    year1_nbc = float(result.monthly_summary["nbc_charge"].sum()) if "nbc_charge" in result.monthly_summary.columns else 0.0
    year1_load_kwh = result.annual_load_kwh
    year1_solar_kwh = result.annual_solar_kwh
    year1_import_kwh = result.annual_import_kwh
    year1_export_kwh = result.annual_export_kwh
    year1_export_peak_kwh = float(result.monthly_summary["export_peak_kwh"].sum())
    year1_export_offpeak_kwh = float(result.monthly_summary["export_offpeak_kwh"].sum())

    # Demand kW columns
    year1_demand_kw_bess = 0.0
    if result_pv_only is not None:
        year1_demand_kw_pv = float(result_pv_only.monthly_summary["peak_demand_kw"].max())
        year1_demand_kw_bess = float(result.monthly_summary["peak_demand_kw"].max())
    else:
        year1_demand_kw_pv = float(result.monthly_summary["peak_demand_kw"].max())

    year1_bill_no_solar = result.annual_bill_without_solar

    # Baseline breakdown (no solar) — use actual no-solar components
    if result.monthly_baseline_details:
        year1_baseline_demand = sum(d.get("demand", 0) for d in result.monthly_baseline_details)
        year1_baseline_energy = sum(d.get("energy", 0) for d in result.monthly_baseline_details)
        year1_baseline_fixed = sum(d.get("fixed", 0) for d in result.monthly_baseline_details)
    else:
        # Fallback: approximate from total no-solar bill.
        # Fixed charges are the same with or without solar.
        # Demand and energy must be split from the remaining no-solar bill.
        # Use the with-solar demand ratio as a rough proxy since no-solar
        # demand would be higher (no solar offset), but we lack the exact value.
        year1_baseline_fixed = year1_fixed
        _remaining = year1_bill_no_solar - year1_baseline_fixed
        if year1_demand + year1_energy > 0:
            _demand_share = year1_demand / (year1_demand + year1_energy)
        else:
            _demand_share = 0.3  # reasonable default split
        year1_baseline_demand = _remaining * _demand_share
        year1_baseline_energy = _remaining * (1.0 - _demand_share)

    rate_mult = rate_escalator_pct / 100.0
    load_mult = load_escalator_pct / 100.0
    degrad_rate = degradation_pct / 100.0

    # Precompute BOTH energy baselines so regime-shift years use the correct
    # energy cost:
    #   raw_energy = sum(import_kwh * energy_rate) across ALL meters — NEM-3/NVBT
    #   tou_energy = positive side of TOU per-period netting — NEM-1/2
    #   tou_credit = negative side of TOU per-period netting — NEM-1/2
    # NOTE: result.raw_annual_energy aggregates ALL meters (including NEM-A
    # aggregated meters), unlike result.hourly_detail which only has the
    # generating meter.
    raw_year1_energy = result.raw_annual_energy
    tou_year1_energy = result.tou_annual_energy
    year1_tou_credit = result.tou_annual_credit

    # Split raw energy into generating-meter vs aggregated-meter portions.
    # For NEM-A, hourly_detail only has the generating meter; the difference
    # is the aggregated meters' raw energy (load-only, no solar).
    # For single-meter, _agg_raw_energy = 0.
    _gen_raw_energy = float(result.hourly_detail["energy_cost"].sum())
    _agg_raw_energy = raw_year1_energy - _gen_raw_energy

    # Blended import rate ($/kWh) for valuing kWh that shift from export→import
    # as solar degrades.  Uses the generating meter's hourly data since only the
    # generating meter has solar (and thus exports that shift to imports).
    # Denominator uses generating meter's import kWh (from hourly_detail),
    # not annual_import_kwh which may include aggregated meters for NEM-A.
    _gen_import_kwh = float(result.hourly_detail["import_kwh"].sum())
    blended_import_rate = (
        _gen_raw_energy / _gen_import_kwh if _gen_import_kwh > 0 else 0.0
    )

    # Multi-year export rates: keyed by calendar year (e.g. {2026: Series, 2027: ...})
    if export_rates_multiyear and len(export_rates_multiyear) >= 1:
        _my_keys = sorted(export_rates_multiyear.keys())
        multiyear_start = _my_keys[0]   # first calendar year in CSV
        multiyear_max = _my_keys[-1]     # last calendar year in CSV
        _export_cagr = _compute_export_cagr(export_rates_multiyear)
    else:
        multiyear_start = 0
        multiyear_max = 0
        _export_cagr = 0.0

    # Multi-year export rates for regime 2
    if export_rates_multiyear_2 and len(export_rates_multiyear_2) >= 1:
        _my2_keys = sorted(export_rates_multiyear_2.keys())
        multiyear_start_2 = _my2_keys[0]
        multiyear_max_2 = _my2_keys[-1]
        _export_cagr_2 = _compute_export_cagr(export_rates_multiyear_2)
    else:
        multiyear_start_2 = 0
        multiyear_max_2 = 0
        _export_cagr_2 = 0.0

    rows = []
    cumulative_savings = 0.0

    for yr in range(1, years + 1):
        active_regime, active_multiyear, active_my_start, active_my_max, active_cagr = (
            _resolve_regime_context(
                yr, nem_regime_1, nem_regime_2, num_years_1,
                export_rates_multiyear, export_rates_multiyear_2,
                multiyear_start, multiyear_max,
                multiyear_start_2, multiyear_max_2,
                _export_cagr, _export_cagr_2,
            )
        )

        row = _compute_year_row(
            yr,
            year1_load_kwh=year1_load_kwh,
            year1_solar_kwh=year1_solar_kwh,
            year1_import_kwh=year1_import_kwh,
            year1_export_kwh=year1_export_kwh,
            year1_export_peak_kwh=year1_export_peak_kwh,
            year1_export_offpeak_kwh=year1_export_offpeak_kwh,
            year1_demand=year1_demand,
            year1_fixed=year1_fixed,
            year1_export=year1_export,
            year1_nbc=year1_nbc,
            year1_demand_kw_pv=year1_demand_kw_pv,
            year1_demand_kw_bess=year1_demand_kw_bess,
            year1_bill_no_solar=year1_bill_no_solar,
            year1_baseline_energy=year1_baseline_energy,
            year1_baseline_demand=year1_baseline_demand,
            year1_baseline_fixed=year1_baseline_fixed,
            tou_year1_energy=tou_year1_energy,
            year1_tou_credit=year1_tou_credit,
            gen_raw_energy=_gen_raw_energy,
            agg_raw_energy=_agg_raw_energy,
            blended_import_rate=blended_import_rate,
            rate_mult=rate_mult,
            load_mult=load_mult,
            degrad_rate=degrad_rate,
            compound_escalation=compound_escalation,
            active_regime=active_regime,
            active_multiyear=active_multiyear,
            active_my_start=active_my_start,
            active_my_max=active_my_max,
            active_cagr=active_cagr,
            cod_year=cod_year,
            nem_regime_1=nem_regime_1,
            nem_regime_2=nem_regime_2,
            num_years_1=num_years_1,
            nbc_rate_2=nbc_rate_2,
            nsc_rate_2=nsc_rate_2,
            result_pv_only=result_pv_only,
            result_hourly_detail=result.hourly_detail,
            result_annual_bill_with_solar=result.annual_bill_with_solar,
            existing_solar_offset_kwh=existing_solar_offset_kwh,
            year1_nsc_adj=result.annual_nsc_adjustment,
            monthly_summary_y1=result.monthly_summary,
            tou_monthly_energy_y1=result.tou_monthly_energy,
            tou_monthly_credit_y1=result.tou_monthly_credit,
            billing_option=result.billing_option,
            min_monthly_charge=result.min_monthly_charge,
        )

        cumulative_savings += row["Annual Savings ($)"]
        row["Cumulative Savings ($)"] = round(cumulative_savings)
        rows.append(row)

    # Rate shift columns (when old-rate baseline is provided)
    if rate_shift_old_baseline is not None:
        _apply_rate_shift_columns(
            rows, rate_shift_old_baseline, rate_mult, load_mult, compound_escalation,
        )

    return pd.DataFrame(rows)


# ─── 38DN brand palette — delegates to modules.ui.tokens ────────────────────
# Kept as module-level aliases so existing code in app.py that imports them
# directly (if any) continues to work. New code should import PALETTE from
# modules.ui.tokens instead.
from modules.ui.tokens import PALETTE as _PAL, PLOTLY_LAYOUT as _PLY_BASE
_38DN_NAVY   = _PAL["navy"]
_38DN_GREEN  = _PAL["green"]
_38DN_TEAL   = _PAL["teal"]
_38DN_BLUE   = _PAL["blue"]
_38DN_AMBER  = _PAL["amber"]
_38DN_GRAY50 = _PAL["slate_50"]
_38DN_INK    = _PAL["ink"]
_38DN_FONT   = "Inter, Aptos Narrow, sans-serif"


def _apply_38dn_layout(fig: go.Figure, *, title: str, x_title: str, y_title: str,
                       height: int = 400, barmode: str | None = None) -> go.Figure:
    """Apply consistent 38DN chart styling via the token-backed base layout
    in :mod:`modules.ui.tokens`. Callers override title / axis labels /
    height / barmode; everything else is uniform across the app.
    """
    # Deep-merge: Plotly's update_layout treats nested dicts positionally,
    # so we spread the base layout then overlay chart-specific overrides.
    fig.update_layout(**_PLY_BASE)
    fig.update_layout(
        title=dict(text=title, font=dict(size=15, color=_38DN_NAVY)),
        xaxis_title=x_title,
        yaxis_title=y_title,
        height=height,
    )
    if barmode:
        fig.update_layout(barmode=barmode)
    return fig


def create_production_vs_load_chart(result: BillingResult) -> go.Figure:
    """Monthly grouped bar chart: Load vs Solar Production with Net Import overlay."""
    df = result.monthly_summary
    fig = go.Figure()
    fig.add_trace(go.Bar(x=MONTH_NAMES, y=df["load_kwh"], name="Load",
                         marker_color=_38DN_NAVY, opacity=0.88))
    fig.add_trace(go.Bar(x=MONTH_NAMES, y=df["solar_kwh"], name="Solar Production",
                         marker_color=_38DN_GREEN, opacity=0.88))
    fig.add_trace(go.Scatter(x=MONTH_NAMES, y=df["import_kwh"], name="Net Import",
                             mode="lines+markers",
                             line=dict(color=_38DN_BLUE, width=2.5),
                             marker=dict(size=8, color=_38DN_BLUE)))
    return _apply_38dn_layout(
        fig, title="Monthly Production vs. Load",
        x_title="Month", y_title="Energy (kWh)", barmode="group",
    )


def create_monthly_bill_chart(result: BillingResult) -> go.Figure:
    """Stacked bar chart of monthly bill components. Charges stack up,
    export credit stacks down (below zero) for visual separation."""
    df = result.monthly_summary
    fig = go.Figure()
    fig.add_trace(go.Bar(x=MONTH_NAMES, y=df["energy_cost"], name="Energy Charges",
                         marker_color=_38DN_NAVY, opacity=0.92))
    fig.add_trace(go.Bar(x=MONTH_NAMES, y=df["total_demand_charge"],
                         name="Demand Charges",
                         marker_color=_38DN_BLUE, opacity=0.92))
    fig.add_trace(go.Bar(x=MONTH_NAMES, y=df["fixed_charge"], name="Fixed Charges",
                         marker_color=_38DN_TEAL, opacity=0.92))
    if "nbc_charge" in df.columns and df["nbc_charge"].sum() > 0:
        fig.add_trace(go.Bar(x=MONTH_NAMES, y=df["nbc_charge"], name="NBC Charges",
                             marker_color=_38DN_AMBER, opacity=0.92))
    fig.add_trace(go.Bar(x=MONTH_NAMES, y=-df["export_credit"], name="Export Credit",
                         marker_color=_38DN_GREEN, opacity=0.92))
    return _apply_38dn_layout(
        fig, title="Monthly Bill Breakdown (With Solar)",
        x_title="Month", y_title="Cost ($)", barmode="relative",
    )


def generate_hourly_csv(result: BillingResult, cod_date=None) -> str:
    """Generate CSV string of hourly detail data for download."""
    df = result.hourly_detail.copy()
    if cod_date is not None:
        df = df[df.index >= pd.Timestamp(cod_date)]
    if "export_kwh" in df.columns and "export_credit" in df.columns:
        exp = df["export_kwh"]
        df["value_of_energy_dollar_per_kwh"] = np.where(
            exp > 0, df["export_credit"] / exp, 0.0,
        )
    df.index.name = "datetime"
    buf = StringIO()
    df.to_csv(buf)
    return buf.getvalue()


def _compute_monthly_export_overrides(
    active_regime, active_multiyear, active_my_start, active_my_max, active_cagr,
    month_tou_credits, rate_factor, volume_ratio,
    hd, _cod_year, yr,
):
    """Compute per-month export credit overrides based on active NEM regime.

    Returns:
        dict[int, float] | None -- per-month override values, or None to use default scaling.
    """
    if active_regime in ("NEM-1", "NEM-2") or active_regime.startswith("NEM-A"):
        # TOU-netted credits scaled by rate escalation and volume change
        month_export_credit_override = {}
        for m in range(1, 13):
            month_export_credit_override[m] = month_tou_credits[m] * rate_factor * volume_ratio
        return month_export_credit_override
    elif active_my_max > 0 and active_multiyear:
        calendar_year = (_cod_year if _cod_year is not None else active_my_start) + (yr - 1)
        cal_yr = max(active_my_start, min(calendar_year, active_my_max))
        yr_rates = active_multiyear[cal_yr].values
        hourly_export = hd["export_kwh"].values
        dt_index = hd.index
        month_idx = dt_index.month
        # NEM-3 export rates come from the CSV — do NOT apply utility
        # rate escalation.  Beyond the CSV range, extrapolate using
        # the 10-year CAGR of the CSV's trailing years.
        _cagr_mult = 1.0
        if calendar_year > active_my_max and active_cagr != 0.0:
            _cagr_mult = (1.0 + active_cagr) ** (calendar_year - active_my_max)
        month_export_credit_override = {}
        for m in range(1, 13):
            mask = month_idx == m
            base = float(np.sum(hourly_export[mask] * yr_rates[mask]))
            month_export_credit_override[m] = base * _cagr_mult * volume_ratio
        return month_export_credit_override
    return None


def _project_single_year_monthly(
    yr, ms, result, result_pv_only, rate_factor, load_factor, solar_factor,
    volume_ratio, import_ratio, _cod_year, _cod_month, _cod_day, cod_date,
    active_regime, _any_nem2,
    month_tou_energy, raw_month_energy, month_wtd_rate,
    month_export_credit_override,
    _any_nsc_regime: bool = False,
):
    """Project one year's 12-month rows for _build_multiyear_monthly_df.

    Returns:
        list[dict] -- one dict per month in this year.
    """
    rows = []
    for _, mrow in ms.iterrows():
        m = int(mrow["month"])

        # Skip pre-COD months for Year 1
        if yr == 1 and cod_date and m < _cod_month:
            continue

        r = {}
        r["Year"] = yr
        if _cod_year is not None:
            r["Calendar Year"] = _cod_year + (yr - 1)
        r["Month"] = MONTH_NAMES[m - 1]

        # Pro-rate COD month (Year 1 only)
        _prorate = 1.0
        if yr == 1 and cod_date and m == _cod_month and _cod_day > 1:
            _days = calendar.monthrange(_cod_year, m)[1]
            _prorate = (_days - _cod_day + 1) / _days
            r["Month"] = f"{MONTH_NAMES[m - 1]} (partial)"

        r["Load (kWh)"] = float(mrow["load_kwh"] * load_factor * _prorate)
        r["Solar (kWh)"] = float(mrow["solar_kwh"] * solar_factor * _prorate)
        r["Import (kWh)"] = float(mrow["import_kwh"] * load_factor * _prorate)
        r["Export (kWh)"] = float(mrow["export_kwh"] * volume_ratio * _prorate)
        r["Export Peak (kWh)"] = float(mrow["export_peak_kwh"] * volume_ratio * _prorate)
        r["Export Off-Peak (kWh)"] = float(mrow["export_offpeak_kwh"] * volume_ratio * _prorate)

        if result_pv_only is not None:
            pv_row = result_pv_only.monthly_summary[result_pv_only.monthly_summary["month"] == m].iloc[0]
            r["Demand kW (PV)"] = round(pv_row["peak_demand_kw"] * load_factor, 2)
            r["Demand kW (PV+BESS)"] = round(mrow["peak_demand_kw"] * load_factor, 2)
        else:
            r["Demand kW (PV)"] = round(mrow["peak_demand_kw"] * load_factor, 2)

        r["Wtd Avg Rate ($/kWh)"] = round(month_wtd_rate[m] * rate_factor, 5)

        # Energy cost depends on active regime:
        #   NEM-1/2 (and NEM-A): TOU-netted energy (exports offset imports within each TOU period)
        #   NEM-3/NVBT: raw import energy cost (no netting; exports valued separately)
        _is_tou_netted = active_regime in ("NEM-1", "NEM-2") or active_regime.startswith("NEM-A")
        if _is_tou_netted:
            r["Energy ($)"] = round(month_tou_energy[m] * load_factor * rate_factor * _prorate, 2)
        else:
            r["Energy ($)"] = round(raw_month_energy[m] * import_ratio * rate_factor * _prorate, 2)
        r["Demand ($)"] = round(mrow["total_demand_charge"] * load_factor * rate_factor, 2)
        r["Fixed ($)"] = round(mrow["fixed_charge"] * rate_factor * _prorate, 2)

        # NBC: only applies during NEM-2 regime years (including NEM-A (NEM-2))
        _m_nbc = 0.0
        if active_regime in ("NEM-2", "NEM-A (NEM-2)") and "nbc_charge" in ms.columns and mrow["nbc_charge"] > 0:
            _m_nbc = round(mrow["nbc_charge"] * rate_factor, 2)
        if _any_nem2:
            r["NBC ($)"] = _m_nbc

        if month_export_credit_override is not None:
            r["Export Credit ($)"] = -round(month_export_credit_override[m] * _prorate, 2)
        else:
            r["Export Credit ($)"] = -round(mrow["export_credit"] * rate_factor * volume_ratio * _prorate, 2)

        # End-of-year NSC clawback lives on month 12 of the Y1 monthly_summary
        # for both NEM-1/2 (retail→NSC) and NEM-3/NBT (avg ACC→NSC).
        # Year-N scaling: NEM-1/2 uses rate × volume; NEM-3 uses volume only
        # (ACC schedule already reflected in Export Credit, not retail rate).
        # Display: rows where the active regime doesn't support NSC show
        # em-dash (NaN) when any year carries NSC; otherwise the column is
        # omitted. Local _m_nsc_adj stays 0.0 so Net Bill arithmetic doesn't
        # get NaN-poisoned.
        _m_nsc_adj = 0.0
        _supports = _supports_nsc(active_regime)
        if _supports and "nsc_adjustment" in ms.columns:
            _y1_m_nsc = float(mrow["nsc_adjustment"])
            if yr == 1:
                _m_nsc_adj = round(_y1_m_nsc * _prorate, 2)
            elif active_regime in _NEM12_REGIMES:
                _m_nsc_adj = round(_y1_m_nsc * rate_factor * volume_ratio, 2)
            else:
                _m_nsc_adj = round(_y1_m_nsc * volume_ratio, 2)
            r["NSC Adj ($)"] = _m_nsc_adj
        elif _any_nsc_regime:
            r["NSC Adj ($)"] = float("nan")

        if yr == 1 and _prorate == 1.0:
            # Year 1: use actual monthly net_bill from billing result
            # (includes min_monthly_charge floors, MBO/ABO credit banking,
            # and NSC true-ups that component reconstruction misses)
            r["Net Bill ($)"] = round(float(mrow["net_bill"]), 2)
        else:
            # Provisional component sum; corrected post-loop via the billing-option
            # simulation when the active regime is NEM-1/2 (gives MBO/ABO floors
            # the right place to clamp). NEM-3 keeps this value because there's
            # no monthly netting / banking — credit just offsets the bill.
            r["Net Bill ($)"] = round(
                r["Energy ($)"] + r["Demand ($)"] + r["Fixed ($)"] + _m_nbc + _m_nsc_adj + r["Export Credit ($)"], 2
            )

        # Baseline bill (no-solar) per month — for Indexed Tariff PPA rate calc
        if result.monthly_baseline_details is not None:
            _bd = result.monthly_baseline_details[m - 1]
            r["Baseline Bill ($)"] = round(
                _bd["energy"] * load_factor * rate_factor * _prorate
                + _bd["demand"] * load_factor * rate_factor * _prorate
                + _bd["fixed"] * rate_factor * _prorate, 2)
        else:
            # Fallback: distribute annual no-solar bill by monthly load share
            _mo_load_share = (
                mrow["load_kwh"] / result.annual_load_kwh
                if result.annual_load_kwh > 0 else 1.0 / 12
            )
            r["Baseline Bill ($)"] = round(
                result.annual_bill_without_solar * _mo_load_share
                * load_factor * rate_factor * _prorate, 2)

        # Monthly solar savings = baseline bill (no-solar) − net bill (with solar).
        # Positive when solar reduces the bill; can go negative in a month if
        # minimum-charge floors or regime transitions flip the balance.
        r["Savings ($)"] = round(r["Baseline Bill ($)"] - r["Net Bill ($)"], 2)

        rows.append(r)

    # Y>1 + NEM-1/2: replace per-month Net Bill with billing-option simulation
    # so MBO credit-banking and min_monthly_charge floors are honored. Without
    # this, gross TOU credit subtracts from gross charges at full magnitude
    # for over-sized PV — which silently understates Y>1 bills.
    _is_tou_netted_year = active_regime in ("NEM-1", "NEM-2") or active_regime.startswith("NEM-A")
    if yr > 1 and len(rows) == 12 and _is_tou_netted_year:
        monthly_components = [
            {
                "energy": r["Energy ($)"],
                # Export Credit ($) is stored as a negative number in the row
                # (it's an offset/credit); the simulator wants positive magnitude.
                "export_credit": -r["Export Credit ($)"],
                "demand": r["Demand ($)"],
                "fixed": r["Fixed ($)"],
                "nbc": r.get("NBC ($)", 0.0),
                "nsc_adj": r.get("NSC Adj ($)", 0.0),
            }
            for r in rows
        ]
        net_bills = simulate_year_under_billing_option(
            monthly_components,
            active_regime,
            result.billing_option,
            result.min_monthly_charge,
        )
        for r, nb in zip(rows, net_bills):
            r["Net Bill ($)"] = round(nb, 2)
            r["Savings ($)"] = round(r["Baseline Bill ($)"] - r["Net Bill ($)"], 2)

    return rows


def _build_multiyear_monthly_df(
    result: BillingResult,
    result_pv_only: BillingResult | None = None,
    rate_escalator_pct: float = 0.0,
    load_escalator_pct: float = 0.0,
    years: int = 1,
    export_rates_multiyear: dict[int, "pd.Series"] | None = None,
    nem_regime_1: str = "NEM-3 / NVBT",
    nem_regime_2: str | None = None,
    num_years_1: int | None = None,
    export_rates_multiyear_2: dict[int, "pd.Series"] | None = None,
    cod_date=None,
    degradation_pct: float = 0.0,
    compound_escalation: bool = True,
) -> pd.DataFrame:
    """Build a multi-year monthly DataFrame (12 × years rows).

    Scales year-1 monthly values with escalators and regime-aware
    export credit / NBC logic matching build_annual_projection.
    """
    _cod_month = cod_date.month if cod_date else 1
    _cod_day = cod_date.day if cod_date else 1
    _cod_year = cod_date.year if cod_date else 2023

    ms = result.monthly_summary
    rate_mult = rate_escalator_pct / 100.0
    load_mult = load_escalator_pct / 100.0
    degrad_rate = degradation_pct / 100.0

    # Determine multi-year export rate calendar-year start (regime 1)
    if export_rates_multiyear and len(export_rates_multiyear) > 1:
        _my_keys = sorted(export_rates_multiyear.keys())
        my_start = _my_keys[0]
        my_max = _my_keys[-1]
    else:
        my_start = 0
        my_max = 0

    # Multi-year export rates for regime 2
    if export_rates_multiyear_2 and len(export_rates_multiyear_2) > 1:
        _my2_keys = sorted(export_rates_multiyear_2.keys())
        my_start_2 = _my2_keys[0]
        my_max_2 = _my2_keys[-1]
    else:
        my_start_2 = 0
        my_max_2 = 0

    # Precompute CAGR for extrapolating export rates beyond CSV range
    if export_rates_multiyear and len(export_rates_multiyear) > 1:
        _my_cagr = _compute_export_cagr(export_rates_multiyear)
    else:
        _my_cagr = 0.0
    if export_rates_multiyear_2 and len(export_rates_multiyear_2) > 1:
        _my_cagr_2 = _compute_export_cagr(export_rates_multiyear_2)
    else:
        _my_cagr_2 = 0.0

    year1_load = float(ms["load_kwh"].sum())
    year1_solar = float(ms["solar_kwh"].sum())
    year1_export = float(ms["export_kwh"].sum())

    # Per-month TOU-netted energy AND credit (for NEM-1/2 regime years)
    # and per-month raw import energy cost (for NEM-3 regime years)
    hd = result.hourly_detail
    # Use pre-computed per-month TOU breakdowns from billing engine (preferred)
    # or fall back to recomputation for legacy BillingResult objects
    if result.tou_monthly_energy is not None and result.tou_monthly_credit is not None:
        month_tou_energy = result.tou_monthly_energy
        month_tou_credits = result.tou_monthly_credit
    else:
        _, _, month_tou_energy, month_tou_credits = _compute_tou_netted_monthly(hd)
    raw_month_energy: dict[int, float] = {}
    month_import_kwh: dict[int, float] = {}
    for month in range(1, 13):
        mm = hd.index.month == month
        raw_month_energy[month] = float(hd.loc[mm, "energy_cost"].sum())
        month_import_kwh[month] = float(hd.loc[mm, "import_kwh"].sum())

    year1_import_total = sum(month_import_kwh.values())

    # Precompute per-month weighted average retail rate (import-weighted)
    month_wtd_rate: dict[int, float] = {}
    for month in range(1, 13):
        mm = hd.index.month == month
        imp = float(hd.loc[mm, "import_kwh"].sum())
        if imp > 0:
            month_wtd_rate[month] = float((hd.loc[mm, "import_kwh"] * hd.loc[mm, "energy_rate"]).sum()) / imp
        else:
            month_wtd_rate[month] = 0.0

    _any_nem2 = any(r in ("NEM-2", "NEM-A (NEM-2)") for r in (nem_regime_1, nem_regime_2) if r)
    _any_nsc_regime = any(_supports_nsc(r) for r in (nem_regime_1, nem_regime_2))

    rows = []
    for yr in range(1, years + 1):
        if compound_escalation:
            rate_factor = (1.0 + rate_mult) ** (yr - 1)
            load_factor = (1.0 + load_mult) ** (yr - 1)
        else:
            rate_factor = 1.0 + rate_mult * (yr - 1)
            load_factor = 1.0 + load_mult * (yr - 1)
        solar_factor = (1.0 - degrad_rate) ** (yr - 1)
        net_delta = year1_load * (load_factor - 1) + year1_solar * (1.0 - solar_factor)
        yr_export_total = max(0, year1_export - net_delta)
        volume_ratio = yr_export_total / year1_export if year1_export > 0 else 0.0

        # Import volume ratio (for scaling raw energy cost under NEM-3)
        absorbed = year1_export - yr_export_total
        yr_import_total = year1_import_total + (net_delta - absorbed)
        import_ratio = yr_import_total / year1_import_total if year1_import_total > 0 else load_factor

        # Determine active regime for this year
        active_regime, active_multiyear, active_my_start, active_my_max, active_cagr = (
            _resolve_regime_context(
                yr, nem_regime_1, nem_regime_2, num_years_1,
                export_rates_multiyear, export_rates_multiyear_2,
                my_start, my_max, my_start_2, my_max_2,
                _my_cagr, _my_cagr_2,
            )
        )

        # Per-month export credit recompute based on active regime
        month_export_credit_override = _compute_monthly_export_overrides(
            active_regime, active_multiyear, active_my_start, active_my_max, active_cagr,
            month_tou_credits, rate_factor, volume_ratio,
            hd, _cod_year, yr,
        )

        yr_rows = _project_single_year_monthly(
            yr, ms, result, result_pv_only, rate_factor, load_factor, solar_factor,
            volume_ratio, import_ratio, _cod_year, _cod_month, _cod_day, cod_date,
            active_regime, _any_nem2,
            month_tou_energy, raw_month_energy, month_wtd_rate,
            month_export_credit_override,
            _any_nsc_regime=_any_nsc_regime,
        )
        rows.extend(yr_rows)

    return pd.DataFrame(rows)


def generate_monthly_csv(
    result: BillingResult,
    result_pv_only: BillingResult | None = None,
    rate_escalator_pct: float = 0.0,
    load_escalator_pct: float = 0.0,
    years: int = 1,
    export_rates_multiyear: dict[int, "pd.Series"] | None = None,
    nem_regime_1: str = "NEM-3 / NVBT",
    nem_regime_2: str | None = None,
    num_years_1: int | None = None,
    export_rates_multiyear_2: dict[int, "pd.Series"] | None = None,
    cod_date=None,
    degradation_pct: float = 0.0,
    compound_escalation: bool = True,
) -> str:
    """Generate CSV string of monthly summary data for download.

    Produces a multi-year monthly table with COD-aware partial months
    and Calendar Year columns.
    """
    df = _build_multiyear_monthly_df(
        result=result,
        result_pv_only=result_pv_only,
        rate_escalator_pct=rate_escalator_pct,
        load_escalator_pct=load_escalator_pct,
        years=max(years, 1),
        export_rates_multiyear=export_rates_multiyear,
        nem_regime_1=nem_regime_1,
        nem_regime_2=nem_regime_2,
        num_years_1=num_years_1,
        export_rates_multiyear_2=export_rates_multiyear_2,
        cod_date=cod_date,
        degradation_pct=degradation_pct,
        compound_escalation=compound_escalation,
    )
    buf = StringIO()
    df.to_csv(buf, index=False)
    return buf.getvalue()


def generate_annual_csv(projection_df: pd.DataFrame) -> str:
    """Generate CSV string of annual projection data for download."""
    buf = StringIO()
    projection_df.to_csv(buf, index=False)
    return buf.getvalue()


def _indexed_tariff_savings_target(
    yr: int,
    base_savings_pct: float,
    savings_escalator_pct: float,
    regime_1_savings_pct: float | None,
    regime_2_savings_pct: float | None,
    nem_regime_2: str | None,
    num_years_1: int | None,
) -> float:
    """Compute the savings target (%) for a given projection year.

    Handles partial specification: if only regime_1 or regime_2 savings
    is set, the set value is used for its regime and falls back to
    regime_1 (or base) for the other.
    """
    # Per-regime savings: activate when any regime-specific value is provided
    if regime_1_savings_pct is not None or regime_2_savings_pct is not None:
        is_regime_2 = bool(nem_regime_2 and num_years_1 and yr > num_years_1)
        if is_regime_2:
            base = (regime_2_savings_pct if regime_2_savings_pct is not None
                    else regime_1_savings_pct if regime_1_savings_pct is not None
                    else base_savings_pct)
            esc_yr = yr - num_years_1
        else:
            base = regime_1_savings_pct if regime_1_savings_pct is not None else base_savings_pct
            esc_yr = yr
        return base + savings_escalator_pct * (esc_yr - 1)
    # Uniform savings across all years
    return base_savings_pct + savings_escalator_pct * (yr - 1)


def build_indexed_tariff_annual(
    annual_proj_df: pd.DataFrame,
    base_savings_pct: float,
    savings_escalator_pct: float = 0.0,
    regime_1_savings_pct: float | None = None,
    regime_2_savings_pct: float | None = None,
    nem_regime_2: str | None = None,
    num_years_1: int | None = None,
    ppa_escalator_pct: float = 0.0,
    ppa_escalator_pct_2: float | None = None,
) -> pd.DataFrame:
    """Build an annual Indexed Tariff table solving for PPA rate per year.

    When ppa_escalator_pct > 0, backsolves a Year 1 PPA rate per NEM regime
    such that the escalated rate delivers the target savings over the regime
    period.  Each year shows the escalated PPA rate and actual savings.

    Without escalator (ppa_escalator_pct == 0), falls back to per-year
    independent backsolve:
        PPA Rate  = Utility Savings × (1 - savings_frac) / Solar kWh
    """
    esc_1 = (ppa_escalator_pct or 0.0) / 100.0
    esc_2 = ((ppa_escalator_pct_2 if ppa_escalator_pct_2 is not None
              else ppa_escalator_pct) or 0.0) / 100.0
    use_escalator = esc_1 > 0 or esc_2 > 0

    # Collect per-year data
    year_data = []
    has_nsc = "NSC Adj ($)" in annual_proj_df.columns
    for _, row in annual_proj_df.iterrows():
        yr = int(row["Year"])
        savings_target = _indexed_tariff_savings_target(
            yr, base_savings_pct, savings_escalator_pct,
            regime_1_savings_pct, regime_2_savings_pct,
            nem_regime_2, num_years_1,
        )
        savings_frac = savings_target / 100.0
        bill_no_solar = row["Bill w/o Solar ($)"]
        bill_solar = row["Bill w/ Solar ($)"]
        solar_kwh = row["Solar (kWh)"]
        utility_savings = bill_no_solar - bill_solar
        cal_yr = int(row["Calendar Year"]) if "Calendar Year" in row.index else None
        nsc_adj = float(row["NSC Adj ($)"]) if has_nsc and pd.notna(row.get("NSC Adj ($)")) else 0.0
        year_data.append((yr, bill_no_solar, bill_solar, solar_kwh,
                          utility_savings, savings_frac, savings_target, cal_yr, nsc_adj))

    # Backsolve Year 1 PPA rate per regime when escalator is set
    yr1_ppa_rate_r1 = 0.0
    yr1_ppa_rate_r2 = 0.0
    if use_escalator:
        # Regime 1 years
        r1_years = [(yr, us, sf, skwh) for yr, _, _, skwh, us, sf, _, _, _ in year_data
                     if not (nem_regime_2 and num_years_1 and yr > num_years_1)]
        if r1_years:
            num = sum(us * (1.0 - sf) for _, us, sf, _ in r1_years)
            den = sum(((1 + esc_1) ** (yr - 1)) * skwh for yr, _, _, skwh in r1_years)
            yr1_ppa_rate_r1 = num / den if den > 0 else 0.0

        # Regime 2 years
        if nem_regime_2 and num_years_1:
            r2_years = [(yr, us, sf, skwh) for yr, _, _, skwh, us, sf, _, _, _ in year_data
                         if yr > num_years_1]
            if r2_years:
                r2_start = num_years_1 + 1
                num = sum(us * (1.0 - sf) for _, us, sf, _ in r2_years)
                den = sum(((1 + esc_2) ** (yr - r2_start)) * skwh for yr, _, _, skwh in r2_years)
                yr1_ppa_rate_r2 = num / den if den > 0 else 0.0

    rows = []
    for yr, bill_no_solar, bill_solar, solar_kwh, utility_savings, savings_frac, savings_target, cal_yr, nsc_adj in year_data:
        if use_escalator:
            # Determine which regime and escalated rate
            if nem_regime_2 and num_years_1 and yr > num_years_1:
                esc = esc_2
                yr1_rate = yr1_ppa_rate_r2
                regime_yr = yr - num_years_1
            else:
                esc = esc_1
                yr1_rate = yr1_ppa_rate_r1
                regime_yr = yr
            ppa_rate = yr1_rate * ((1 + esc) ** (regime_yr - 1))
            customer_savings = utility_savings - ppa_rate * solar_kwh
        else:
            # No escalator: per-year independent backsolve
            if solar_kwh > 0:
                ppa_rate = utility_savings * (1.0 - savings_frac) / solar_kwh
            else:
                ppa_rate = 0.0
            customer_savings = utility_savings * savings_frac

        r = {"Year": yr}
        if cal_yr is not None:
            r["Calendar Year"] = cal_yr
        r.update({
            "Bill w/o Solar ($)": round(bill_no_solar, 2),
            "Bill w/ Solar ($)": round(bill_solar, 2),
            "Utility Savings ($)": round(utility_savings, 2),
            "Customer Savings ($)": round(customer_savings, 2),
            "Solar (kWh)": round(solar_kwh, 0),
            "Savings Target (%)": round(savings_target, 1),
            "PPA Rate ($/kWh)": round(ppa_rate, 5),
        })
        if has_nsc:
            r["NSC Adj ($)"] = round(nsc_adj, 2)
        rows.append(r)

    return pd.DataFrame(rows)


def build_indexed_tariff_monthly(
    multiyear_monthly_df: pd.DataFrame,
    base_savings_pct: float,
    savings_escalator_pct: float = 0.0,
    regime_1_savings_pct: float | None = None,
    regime_2_savings_pct: float | None = None,
    nem_regime_2: str | None = None,
    num_years_1: int | None = None,
    ppa_escalator_pct: float = 0.0,
    ppa_escalator_pct_2: float | None = None,
) -> pd.DataFrame:
    """Build a monthly Indexed Tariff table solving for PPA rate per month.

    When ppa_escalator_pct > 0, uses the annual Year 1 PPA rate (backsolve)
    escalated to each year.  The same rate applies to all months within a year.
    """
    esc_1 = (ppa_escalator_pct or 0.0) / 100.0
    esc_2 = ((ppa_escalator_pct_2 if ppa_escalator_pct_2 is not None
              else ppa_escalator_pct) or 0.0) / 100.0
    use_escalator = esc_1 > 0 or esc_2 > 0

    # Collect per-row data for escalator backsolve
    row_data = []
    has_nsc = "NSC Adj ($)" in multiyear_monthly_df.columns
    for _, row in multiyear_monthly_df.iterrows():
        yr = int(row["Year"])
        savings_target = _indexed_tariff_savings_target(
            yr, base_savings_pct, savings_escalator_pct,
            regime_1_savings_pct, regime_2_savings_pct,
            nem_regime_2, num_years_1,
        )
        savings_frac = savings_target / 100.0
        baseline_bill = row.get("Baseline Bill ($)", 0.0)
        if baseline_bill is None or (isinstance(baseline_bill, float) and np.isnan(baseline_bill)):
            baseline_bill = 0.0
        net_bill = row["Net Bill ($)"]
        solar_kwh = row["Solar (kWh)"]
        utility_savings = baseline_bill - net_bill
        cal_yr = int(row["Calendar Year"]) if "Calendar Year" in row.index else None
        month = row["Month"]
        nsc_adj = float(row["NSC Adj ($)"]) if has_nsc and pd.notna(row.get("NSC Adj ($)")) else 0.0
        row_data.append((yr, month, baseline_bill, net_bill, solar_kwh,
                         utility_savings, savings_frac, savings_target, cal_yr, nsc_adj))

    # Backsolve Year 1 PPA rates per regime when escalator is set
    yr1_ppa_rate_r1 = 0.0
    yr1_ppa_rate_r2 = 0.0
    if use_escalator:
        # Regime 1 months
        r1_rows = [(yr, us, sf, skwh) for yr, _, _, _, skwh, us, sf, _, _, _ in row_data
                    if not (nem_regime_2 and num_years_1 and yr > num_years_1)]
        if r1_rows:
            num = sum(us * (1.0 - sf) for _, us, sf, _ in r1_rows)
            den = sum(((1 + esc_1) ** (yr - 1)) * skwh for yr, _, _, skwh in r1_rows)
            yr1_ppa_rate_r1 = num / den if den > 0 else 0.0

        # Regime 2 months
        if nem_regime_2 and num_years_1:
            r2_rows = [(yr, us, sf, skwh) for yr, _, _, _, skwh, us, sf, _, _, _ in row_data
                        if yr > num_years_1]
            if r2_rows:
                r2_start = num_years_1 + 1
                num = sum(us * (1.0 - sf) for _, us, sf, _ in r2_rows)
                den = sum(((1 + esc_2) ** (yr - r2_start)) * skwh for yr, _, _, skwh in r2_rows)
                yr1_ppa_rate_r2 = num / den if den > 0 else 0.0

    rows = []
    for yr, month, baseline_bill, net_bill, solar_kwh, utility_savings, savings_frac, savings_target, cal_yr, nsc_adj in row_data:
        if use_escalator:
            if nem_regime_2 and num_years_1 and yr > num_years_1:
                esc = esc_2
                yr1_rate = yr1_ppa_rate_r2
                regime_yr = yr - num_years_1
            else:
                esc = esc_1
                yr1_rate = yr1_ppa_rate_r1
                regime_yr = yr
            ppa_rate = yr1_rate * ((1 + esc) ** (regime_yr - 1))
            customer_savings = utility_savings - ppa_rate * solar_kwh
        else:
            if solar_kwh > 0:
                ppa_rate = utility_savings * (1.0 - savings_frac) / solar_kwh
            else:
                ppa_rate = 0.0
            customer_savings = utility_savings * savings_frac

        r = {"Year": yr}
        if cal_yr is not None:
            r["Calendar Year"] = cal_yr
        r.update({
            "Month": month,
            "Bill w/o Solar ($)": round(baseline_bill, 2),
            "Net Bill ($)": round(net_bill, 2),
            "Utility Savings ($)": round(utility_savings, 2),
            "Customer Savings ($)": round(customer_savings, 2),
            "Solar (kWh)": round(solar_kwh, 0),
            "Savings Target (%)": round(savings_target, 1),
            "PPA Rate ($/kWh)": round(ppa_rate, 5),
        })
        if has_nsc:
            r["NSC Adj ($)"] = round(nsc_adj, 2)
        rows.append(r)

    return pd.DataFrame(rows)


def build_grid_exchange_summary(
    result: BillingResult, peak_period_idx: int | frozenset[int] = 0,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build monthly grid import/export summary broken out by peak vs off-peak.

    Args:
        result: BillingResult with hourly_detail containing energy_period column
        peak_period_idx: TOU period index(es) considered "peak" (highest rate).
            Can be a single int or a frozenset of ints.

    Returns:
        (display_df, raw_df) — formatted DataFrame with TOTAL row, and raw numeric DataFrame
    """
    hd = result.hourly_detail
    hd_month = hd.index.month
    _has_bess = "batt_to_grid_kwh" in hd.columns

    ep = hd["energy_period"].values
    if isinstance(peak_period_idx, (set, frozenset)):
        is_peak = np.isin(ep, list(peak_period_idx))
    else:
        is_peak = ep == peak_period_idx

    rows = []
    for m in range(1, 13):
        mm = hd_month == m
        peak_m = mm & is_peak
        offpeak_m = mm & ~is_peak

        imp_peak = float(hd.loc[peak_m, "import_kwh"].sum())
        imp_offpeak = float(hd.loc[offpeak_m, "import_kwh"].sum())
        imp_total = imp_peak + imp_offpeak

        exp_peak = float(hd.loc[peak_m, "export_kwh"].sum())
        exp_offpeak = float(hd.loc[offpeak_m, "export_kwh"].sum())
        exp_total = exp_peak + exp_offpeak

        cost_peak = float(hd.loc[peak_m, "energy_cost"].sum())
        cost_offpeak = float(hd.loc[offpeak_m, "energy_cost"].sum())
        cost_total = cost_peak + cost_offpeak

        credit_peak = float(hd.loc[peak_m, "export_credit"].sum())
        credit_offpeak = float(hd.loc[offpeak_m, "export_credit"].sum())
        credit_total = credit_peak + credit_offpeak

        solar_m = float(hd.loc[mm, "solar_kwh"].sum())

        row = {
            "Month": MONTH_NAMES[m - 1],
            "Degraded Solar- Assumed Demand Offset (kWh)": round(solar_m, 0),
            "Import Total (kWh)": round(imp_total, 0),
            "↳ Peak (kWh)": round(imp_peak, 0),
            "↳ Off-Peak (kWh)": round(imp_offpeak, 0),
            "Export Total (kWh)": round(exp_total, 0),
        }
        if _has_bess:
            bess_exp = float(hd.loc[mm, "batt_to_grid_kwh"].sum())
            pv_exp = max(0.0, exp_total - bess_exp)
            row["↳ PV (kWh)"] = round(pv_exp, 0)
            row["↳ BESS (kWh)"] = round(bess_exp, 0)
        # Export totals and sub-components — sub-cols labelled with a trailing
        # column-specific disambiguator because the Import block above already
        # uses "↳ Peak (kWh)" / "↳ Off-Peak (kWh)" as its sub-cols.
        row.update({
            "↳ Export Peak (kWh)": round(exp_peak, 0),
            "↳ Export Off-Peak (kWh)": round(exp_offpeak, 0),
            # Import cost = outflow → negated so it renders red accounting-negative.
            "Import Cost Total ($)": -round(cost_total, 0),
            "↳ Cost Peak ($)": -round(cost_peak, 0),
            "↳ Cost Off-Peak ($)": -round(cost_offpeak, 0),
            # Export credit = inflow → stays positive.
            "Export Credit Total ($)": round(credit_total, 0),
            "↳ Credit Peak ($)": round(credit_peak, 0),
            "↳ Credit Off-Peak ($)": round(credit_offpeak, 0),
        })
        rows.append(row)

    raw_df = pd.DataFrame(rows)

    # Format display copy
    df = raw_df.copy()
    for c in [col for col in df.columns if "(kWh)" in col]:
        df[c] = df[c].apply(fmt_num)
    for c in [col for col in df.columns if "($)" in col]:
        df[c] = df[c].apply(fmt_dollar)

    # TOTAL row
    totals = {"Month": "TOTAL"}
    for c in raw_df.columns:
        if c == "Month":
            continue
        if "(kWh)" in c:
            totals[c] = fmt_num(raw_df[c].sum())
        elif "($)" in c:
            totals[c] = fmt_dollar(raw_df[c].sum())
    df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)

    return df, raw_df


def build_battery_kpi_summary(
    result_pv_only: BillingResult,
    result_batt: BillingResult,
    capacity_kwh: float,
) -> dict:
    """Compute battery-specific KPIs comparing PV-only to PV+battery results.

    Args:
        result_pv_only: BillingResult from PV-only simulation
        result_batt: BillingResult from PV+battery simulation
        capacity_kwh: Battery nameplate capacity (kWh)

    Returns:
        dict of KPI name -> value
    """
    hd = result_batt.hourly_detail

    # --- Charge / discharge totals ---
    total_charge = float(hd["batt_charge_kwh"].sum()) if "batt_charge_kwh" in hd.columns else 0.0
    total_discharge_to_load = float(hd["batt_to_load_kwh"].sum()) if "batt_to_load_kwh" in hd.columns else 0.0
    total_discharge_to_grid = float(hd["batt_to_grid_kwh"].sum()) if "batt_to_grid_kwh" in hd.columns else 0.0
    total_discharge = total_discharge_to_load + total_discharge_to_grid

    # --- Cycles estimate: throughput / (2 * capacity) ---
    throughput = total_charge + total_discharge
    cycles = throughput / (2.0 * capacity_kwh) if capacity_kwh > 0 else 0.0

    # --- PV self-consumption ---
    # Self-consumed PV = solar production - export
    pv_only_self_consumption = result_pv_only.annual_solar_kwh - result_pv_only.annual_export_kwh
    batt_self_consumption = result_batt.annual_solar_kwh - result_batt.annual_export_kwh

    pv_only_self_pct = (
        pv_only_self_consumption / result_pv_only.annual_solar_kwh * 100
        if result_pv_only.annual_solar_kwh > 0 else 0.0
    )
    batt_self_pct = (
        batt_self_consumption / result_batt.annual_solar_kwh * 100
        if result_batt.annual_solar_kwh > 0 else 0.0
    )
    self_consumption_increase_pct = batt_self_pct - pv_only_self_pct

    # --- Export change ---
    export_change_kwh = result_batt.annual_export_kwh - result_pv_only.annual_export_kwh
    export_change_pct = (
        export_change_kwh / result_pv_only.annual_export_kwh * 100
        if result_pv_only.annual_export_kwh > 0 else 0.0
    )

    # --- Peak demand before vs after ---
    pv_only_peak = float(result_pv_only.monthly_summary["peak_demand_kw"].max())
    batt_peak = float(result_batt.monthly_summary["peak_demand_kw"].max())
    peak_reduction_kw = pv_only_peak - batt_peak
    peak_reduction_pct = (
        peak_reduction_kw / pv_only_peak * 100 if pv_only_peak > 0 else 0.0
    )

    # --- Import change ---
    import_change_kwh = result_batt.annual_import_kwh - result_pv_only.annual_import_kwh

    return {
        "total_charge_kwh": round(total_charge, 1),
        "total_discharge_kwh": round(total_discharge, 1),
        "discharge_to_load_kwh": round(total_discharge_to_load, 1),
        "discharge_to_grid_kwh": round(total_discharge_to_grid, 1),
        "throughput_kwh": round(throughput, 1),
        "cycles": round(cycles, 1),
        "pv_self_consumption_pv_only_pct": round(pv_only_self_pct, 1),
        "pv_self_consumption_batt_pct": round(batt_self_pct, 1),
        "self_consumption_increase_pct": round(self_consumption_increase_pct, 1),
        "export_change_kwh": round(export_change_kwh, 1),
        "export_change_pct": round(export_change_pct, 1),
        "pv_only_peak_kw": round(pv_only_peak, 2),
        "batt_peak_kw": round(batt_peak, 2),
        "peak_reduction_kw": round(peak_reduction_kw, 2),
        "peak_reduction_pct": round(peak_reduction_pct, 1),
        "import_change_kwh": round(import_change_kwh, 1),
    }


def _build_summary_df(
    sim_name, system_size_kw, dc_ac_ratio, production_summary,
    location_input, lat, lon, system_life_years,
    nem_regime_1, nem_regime_2, num_years_1,
    battery_capacity_kwh, discharge_limit_pct,
    utility_name, selected_rate_name,
    rate_escalator_pct, load_escalator_pct,
    result, cod_date,
) -> tuple[pd.DataFrame, list[tuple]]:
    """Build the Summary sheet DataFrame and raw rows list."""
    annual_solar = result.annual_solar_kwh
    if production_summary and "ac_annual" in production_summary:
        annual_production = production_summary["ac_annual"]
    else:
        annual_production = annual_solar

    yield_kwh_kw = (
        round(annual_production / system_size_kw, 1)
        if system_size_kw > 0 else 0.0
    )

    system_size_kwac = round(system_size_kw / dc_ac_ratio, 2) if dc_ac_ratio > 0 else system_size_kw

    self_consumed = annual_solar - result.annual_export_kwh
    self_consumption_frac = (
        self_consumed / annual_solar
        if annual_solar > 0 else 0.0
    )
    export_frac = 1.0 - self_consumption_frac

    regime_2_term = (
        system_life_years - num_years_1
        if num_years_1 is not None else None
    )

    summary_rows = [
        ("Simulation Name", sim_name or "N/A"),
        ("Commercial Operation Date", cod_date.strftime("%B %d, %Y") if cod_date else "N/A"),
        ("System Size (kW-DC)", round(system_size_kw, 2)),
        ("System Size (kW-AC)", system_size_kwac),
        ("Yield (kWh/kW)", yield_kwh_kw),
        ("Annual Production (kWh)", round(annual_production, 0)),
        ("Self-Consumption (%)", self_consumption_frac),
        ("Export (%)", export_frac),
        ("System Life (years)", system_life_years),
        ("Location", location_input or "N/A"),
        ("Latitude", round(lat, 4) if lat is not None else "N/A"),
        ("Longitude", round(lon, 4) if lon is not None else "N/A"),
        ("NEM Regime 1", nem_regime_1),
        ("NEM Regime 1 Term (years)", num_years_1 if num_years_1 is not None else system_life_years),
        ("NEM Regime 2", nem_regime_2 or "N/A"),
        ("NEM Regime 2 Term (years)", regime_2_term if regime_2_term is not None else "N/A"),
        ("BESS Size (kWh)", battery_capacity_kwh),
        ("BESS Export Limit (%)", round(discharge_limit_pct * 100, 1) if discharge_limit_pct else 0.0),
        ("Utility", utility_name or "N/A"),
        ("Rate Tariff", selected_rate_name or "N/A"),
        ("Utility Escalator (%/yr)", rate_escalator_pct),
        ("Demand Escalator (%/yr)", load_escalator_pct),
    ]
    # Rate shift analysis rows
    if result.old_rate_annual_baseline is not None:
        summary_rows.append(("Old Rate Annual Baseline ($)", round(result.old_rate_annual_baseline, 2)))
    if result.rate_shift_annual_savings is not None:
        summary_rows.append(("Rate Shift Annual Savings ($)", round(result.rate_shift_annual_savings, 2)))
        total_savings = result.annual_savings + result.rate_shift_annual_savings
        summary_rows.append(("Total Combined Savings ($)", round(total_savings, 2)))
    summary_df = pd.DataFrame(summary_rows, columns=["Parameter", "Value"])
    return summary_df, summary_rows


def _build_hourly_sheets(
    result, battery_capacity_kwh, nem_regime_1, export_rates_8760,
) -> tuple[pd.DataFrame, pd.DataFrame, bool, np.ndarray, np.ndarray, np.ndarray]:
    """Build Export Rates (Hourly) and Retail Rates (Hourly) DataFrames.

    Returns:
        (export_hourly_df, retail_hourly_df, _has_bess, exp_kwh, exp_credit, _safe_exp)
    """
    hd = result.hourly_detail
    exp_kwh = hd["export_kwh"].values
    exp_credit = np.abs(hd["export_credit"].values)
    _has_bess = battery_capacity_kwh > 0 and "batt_to_grid_kwh" in hd.columns

    # Build columns in desired order: Generation, Export, [PV, BESS], Rate, VoE, [VoE PV, VoE BESS]
    export_hourly_data = {"Datetime": hd.index}
    export_hourly_data["Generation (kWh)"] = hd["solar_kwh"].values
    export_hourly_data["Export (kWh)"] = exp_kwh

    if _has_bess:
        bess_exp = hd["batt_to_grid_kwh"].values
        pv_exp = np.maximum(0.0, exp_kwh - bess_exp)
        export_hourly_data["Export PV (kWh)"] = pv_exp
        export_hourly_data["Export BESS (kWh)"] = bess_exp

    # Rate column: NEM-1/2 -> retail TOU rate; NEM-3/NVBT -> ACC export rate
    if nem_regime_1 in ("NEM-1", "NEM-2") or export_rates_8760 is None:
        export_hourly_data["Export Rate ($/kWh)"] = hd["energy_rate"].values
    else:
        export_hourly_data["Export Rate ($/kWh)"] = (
            export_rates_8760.values
            if hasattr(export_rates_8760, "values") else export_rates_8760
        )

    # Value of Energy: |export_credit| / export_kwh (0 when no export)
    _safe_exp = np.where(exp_kwh > 0, exp_kwh, 1.0)
    export_hourly_data["Value of Energy ($/kWh)"] = np.where(
        exp_kwh > 0, exp_credit / _safe_exp, 0.0,
    )

    if _has_bess:
        hourly_rate = np.where(exp_kwh > 0, exp_credit / _safe_exp, 0.0)
        pv_credit = pv_exp * hourly_rate
        bess_credit = bess_exp * hourly_rate
        _safe_pv = np.where(pv_exp > 0, pv_exp, 1.0)
        _safe_bess = np.where(bess_exp > 0, bess_exp, 1.0)
        export_hourly_data["Value of Energy PV ($/kWh)"] = np.where(
            pv_exp > 0, pv_credit / _safe_pv, 0.0,
        )
        export_hourly_data["Value of Energy BESS ($/kWh)"] = np.where(
            bess_exp > 0, bess_credit / _safe_bess, 0.0,
        )
    export_hourly_df = pd.DataFrame(export_hourly_data)

    retail_hourly_df = pd.DataFrame({
        "Datetime": hd.index,
        "Retail Rate ($/kWh)": hd["energy_rate"].values,
        "Import (kWh)": hd["import_kwh"].values,
    })

    return export_hourly_df, retail_hourly_df, _has_bess, exp_kwh, exp_credit, _safe_exp


def _build_monthly_sheets(
    result, result_pv_only, rate_escalator_pct, load_escalator_pct,
    years, export_rates_multiyear, nem_regime_1, nem_regime_2,
    num_years_1, export_rates_multiyear_2, cod_date, degradation_pct,
    _has_bess, exp_kwh, exp_credit, _safe_exp,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Build Export Rates (Monthly) and Retail Rates (Monthly) DataFrames.

    Returns:
        (export_monthly_df, retail_monthly_df)
    """
    hd = result.hourly_detail
    monthly_df = _build_multiyear_monthly_df(
        result=result,
        result_pv_only=result_pv_only,
        rate_escalator_pct=rate_escalator_pct,
        load_escalator_pct=load_escalator_pct,
        years=years,
        export_rates_multiyear=export_rates_multiyear,
        nem_regime_1=nem_regime_1,
        nem_regime_2=nem_regime_2,
        num_years_1=num_years_1,
        export_rates_multiyear_2=export_rates_multiyear_2,
        cod_date=cod_date,
        degradation_pct=degradation_pct,
    )
    export_monthly_cols = ["Year", "Calendar Year", "Month", "Solar (kWh)", "Export (kWh)", "Export Credit ($)", "NSC Adj ($)"]
    retail_monthly_cols = ["Year", "Calendar Year", "Month", "Import (kWh)", "Wtd Avg Rate ($/kWh)", "Energy ($)"]
    export_monthly_df = monthly_df[[c for c in export_monthly_cols if c in monthly_df.columns]].copy()
    if "Solar (kWh)" in export_monthly_df.columns:
        export_monthly_df.rename(columns={"Solar (kWh)": "Generation (kWh)"}, inplace=True)
    # Weighted average Value of Energy: |Export Credit| / Export kWh per month
    if "Export (kWh)" in export_monthly_df.columns and "Export Credit ($)" in export_monthly_df.columns:
        m_exp = export_monthly_df["Export (kWh)"]
        m_credit = export_monthly_df["Export Credit ($)"].abs()
        _safe_m_exp = np.where(m_exp > 0, m_exp, 1.0)
        export_monthly_df["Value of Energy ($/kWh)"] = np.where(
            m_exp > 0, m_credit / _safe_m_exp, 0.0,
        )
    # PV vs BESS monthly export breakdown (PV+BESS projects only)
    if _has_bess:
        # Aggregate PV/BESS exports and credits per month from hourly data
        hourly_rate = np.where(exp_kwh > 0, exp_credit / _safe_exp, 0.0)
        bess_exp_h = hd["batt_to_grid_kwh"].values
        pv_exp_h = np.maximum(0.0, exp_kwh - bess_exp_h)
        pv_credit_h = pv_exp_h * hourly_rate
        bess_credit_h = bess_exp_h * hourly_rate
        month_idx = hd.index.month
        _pv_exp_mo, _bess_exp_mo = {}, {}
        _pv_cred_mo, _bess_cred_mo = {}, {}
        for m in range(1, 13):
            mm = month_idx == m
            _pv_exp_mo[m] = float(pv_exp_h[mm].sum())
            _bess_exp_mo[m] = float(bess_exp_h[mm].sum())
            _pv_cred_mo[m] = float(pv_credit_h[mm].sum())
            _bess_cred_mo[m] = float(bess_credit_h[mm].sum())
        # Map Year-1 monthly values into multi-year rows (scale by volume ratio)
        if "Month" in export_monthly_df.columns:
            _mo_map = {name: idx + 1 for idx, name in enumerate(MONTH_NAMES)}
            row_months = export_monthly_df["Month"].map(
                lambda x: _mo_map.get(x.split(" ")[0] if isinstance(x, str) else x, 0)
            )
            yr1_total_exp = sum(_pv_exp_mo[m] + _bess_exp_mo[m] for m in range(1, 13))
            row_total = export_monthly_df["Export (kWh)"].values
            pv_exp_col, bess_exp_col = [], []
            pv_voe_col, bess_voe_col = [], []
            for i, m in enumerate(row_months):
                if m == 0 or yr1_total_exp == 0:
                    pv_exp_col.append(0.0)
                    bess_exp_col.append(0.0)
                    pv_voe_col.append(0.0)
                    bess_voe_col.append(0.0)
                    continue
                yr1_mo_total = _pv_exp_mo[m] + _bess_exp_mo[m]
                if yr1_mo_total > 0:
                    pv_frac = _pv_exp_mo[m] / yr1_mo_total
                    bess_frac = _bess_exp_mo[m] / yr1_mo_total
                else:
                    pv_frac, bess_frac = 0.0, 0.0
                mo_total = row_total[i]
                pv_e = mo_total * pv_frac
                bess_e = mo_total * bess_frac
                pv_exp_col.append(round(pv_e, 1))
                bess_exp_col.append(round(bess_e, 1))
                # But VoE should reflect the actual rate each component earns,
                # not just the average. Use year-1 credit ratios.
                if _pv_exp_mo[m] > 0 and _bess_exp_mo[m] > 0:
                    pv_rate_yr1 = _pv_cred_mo[m] / _pv_exp_mo[m]
                    bess_rate_yr1 = _bess_cred_mo[m] / _bess_exp_mo[m]
                elif _pv_exp_mo[m] > 0:
                    pv_rate_yr1 = _pv_cred_mo[m] / _pv_exp_mo[m]
                    bess_rate_yr1 = 0.0
                elif _bess_exp_mo[m] > 0:
                    pv_rate_yr1 = 0.0
                    bess_rate_yr1 = _bess_cred_mo[m] / _bess_exp_mo[m]
                else:
                    pv_rate_yr1 = 0.0
                    bess_rate_yr1 = 0.0
                # Scale rates by the same factor as overall VoE
                mo_voe = export_monthly_df["Value of Energy ($/kWh)"].iloc[i] if mo_total > 0 else 0.0
                yr1_voe = (_pv_cred_mo[m] + _bess_cred_mo[m]) / yr1_mo_total if yr1_mo_total > 0 else 0.0
                rate_scale = mo_voe / yr1_voe if yr1_voe > 0 else 1.0
                pv_voe_col.append(round(pv_rate_yr1 * rate_scale, 5))
                bess_voe_col.append(round(bess_rate_yr1 * rate_scale, 5))
            export_monthly_df["Export PV (kWh)"] = pv_exp_col
            export_monthly_df["Export BESS (kWh)"] = bess_exp_col
            export_monthly_df["Value of Energy PV ($/kWh)"] = pv_voe_col
            export_monthly_df["Value of Energy BESS ($/kWh)"] = bess_voe_col
            # Reorder so Export PV/BESS columns follow Export (kWh)
            _ordered = [c for c in export_monthly_df.columns if c not in ("Export PV (kWh)", "Export BESS (kWh)")]
            _insert_idx = _ordered.index("Export (kWh)") + 1
            _ordered = _ordered[:_insert_idx] + ["Export PV (kWh)", "Export BESS (kWh)"] + _ordered[_insert_idx:]
            export_monthly_df = export_monthly_df[_ordered]
    retail_monthly_df = monthly_df[[c for c in retail_monthly_cols if c in monthly_df.columns]].copy()

    return export_monthly_df, retail_monthly_df


def _build_projection_sheet(export_monthly_df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate monthly export data into per-year rows for Export Rates (Annual).

    Returns:
        export_annual_df
    """
    _ann_rows = []
    if "Year" in export_monthly_df.columns:
        for yr_val in export_monthly_df["Year"].unique():
            yr_slice = export_monthly_df[export_monthly_df["Year"] == yr_val]
            row: dict = {"Year": int(yr_val)}
            if "Calendar Year" in yr_slice.columns:
                row["Calendar Year"] = int(yr_slice["Calendar Year"].iloc[0])
            # Sum kWh and $ columns
            for col in ["Generation (kWh)", "Export (kWh)", "Export PV (kWh)",
                         "Export BESS (kWh)", "Export Credit ($)", "NSC Adj ($)"]:
                if col in yr_slice.columns:
                    row[col] = round(float(yr_slice[col].sum()), 1)
            # Weighted average Value of Energy = |credit| / export_kwh
            yr_exp = row.get("Export (kWh)", 0.0)
            yr_cred = abs(row.get("Export Credit ($)", 0.0))
            row["Value of Energy ($/kWh)"] = round(yr_cred / yr_exp, 5) if yr_exp > 0 else 0.0
            # BESS VoE columns
            if "Export PV (kWh)" in row and "Export BESS (kWh)" in row:
                for comp, exp_key in [("PV", "Export PV (kWh)"), ("BESS", "Export BESS (kWh)")]:
                    voe_col = f"Value of Energy {comp} ($/kWh)"
                    if voe_col in yr_slice.columns:
                        c_exp = row.get(exp_key, 0.0)
                        # Weighted avg VoE for component = sum(kWh_i * VoE_i) / sum(kWh_i)
                        comp_exp_col = yr_slice[exp_key]
                        comp_voe_col = yr_slice[voe_col]
                        weighted_sum = float((comp_exp_col * comp_voe_col).sum())
                        row[voe_col] = round(weighted_sum / c_exp, 5) if c_exp > 0 else 0.0
            _ann_rows.append(row)
    export_annual_df = pd.DataFrame(_ann_rows)
    # Match column order to monthly sheet (minus Month)
    _ann_col_order = [c for c in export_monthly_df.columns if c != "Month" and c in export_annual_df.columns]
    export_annual_df = export_annual_df[[c for c in _ann_col_order if c in export_annual_df.columns]]
    return export_annual_df


def _write_excel_workbook(
    summary_df, summary_rows, annual_projection_df,
    export_hourly_df, export_monthly_df, export_annual_df,
    retail_hourly_df, retail_monthly_df,
) -> bytes:
    """Assemble and format the multi-sheet Excel workbook.

    Returns bytes of the .xlsx file content.
    """
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        annual_display_df = _negate_outflow_columns(annual_projection_df)
        annual_display_df.to_excel(writer, sheet_name="Annual Savings", index=False)
        export_hourly_df.to_excel(writer, sheet_name="Export Rates (Hourly)", index=False)
        export_monthly_df.to_excel(writer, sheet_name="Export Rates (Monthly)", index=False)
        export_annual_df.to_excel(writer, sheet_name="Export Rates (Annual)", index=False)
        retail_hourly_df.to_excel(writer, sheet_name="Retail Rates (Hourly)", index=False)
        retail_monthly_df.to_excel(writer, sheet_name="Retail Rates (Monthly)", index=False)

        # Style header rows (row 1) across all sheets
        from openpyxl.styles import Alignment, Font, PatternFill
        _header_fill = PatternFill(start_color="0E2841", end_color="0E2841", fill_type="solid")
        _header_font = Font(color="FFFFFF", bold=True)
        for ws in writer.sheets.values():
            for cell in ws[1]:
                cell.fill = _header_fill
                cell.font = _header_font

        # Left-align the Value column on the Summary sheet
        ws_summary = writer.sheets["Summary"]
        _left = Alignment(horizontal="left")
        for row_idx in range(1, len(summary_df) + 2):  # header + data rows
            ws_summary.cell(row=row_idx, column=2).alignment = _left

        # 38DN number formats — see EXCEL_FMT_* at module top. Four-section
        # accounting layout so $0 / 0 kWh / $0.00000 render as en-dash.
        _fmt_kwh = EXCEL_FMT_KWH
        _fmt_dollar = EXCEL_FMT_DOLLAR
        _fmt_dollar_acct = EXCEL_FMT_DOLLAR_ACCT
        _fmt_rate = EXCEL_FMT_RATE
        _fmt_pct = EXCEL_FMT_PCT

        # Format percentage rows on the Summary sheet
        _pct_params = {"Self-Consumption (%)", "Export (%)"}
        for row_idx, (param, _) in enumerate(summary_rows, start=2):
            if param in _pct_params:
                ws_summary.cell(row=row_idx, column=2).number_format = _fmt_pct

        for sheet_name, df, dollar_fmt in [
            ("Export Rates (Hourly)", export_hourly_df, _fmt_dollar),
            ("Export Rates (Monthly)", export_monthly_df, _fmt_dollar),
            ("Export Rates (Annual)", export_annual_df, _fmt_dollar),
            ("Retail Rates (Hourly)", retail_hourly_df, _fmt_dollar),
            ("Retail Rates (Monthly)", retail_monthly_df, _fmt_dollar),
            ("Annual Savings", annual_display_df, _fmt_dollar_acct),
        ]:
            ws = writer.sheets[sheet_name]
            for col_idx, col_name in enumerate(df.columns, start=1):
                if "(kWh)" in col_name:
                    fmt = _fmt_kwh
                elif "($/kWh)" in col_name:
                    fmt = _fmt_rate
                elif "($)" in col_name:
                    fmt = dollar_fmt
                elif "kW" in col_name and "(kWh)" not in col_name:
                    fmt = _fmt_kwh
                else:
                    continue
                for row_idx in range(2, len(df) + 2):  # skip header row
                    ws.cell(row=row_idx, column=col_idx).number_format = fmt

        # Auto-fit column widths across all sheets
        from openpyxl.utils import get_column_letter
        for ws in writer.sheets.values():
            for col_idx in range(1, ws.max_column + 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for row_idx in range(1, min(ws.max_row + 1, 1002)):  # sample header + up to 1000 rows
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell_len = len(str(cell.value))
                        if cell_len > max_len:
                            max_len = cell_len
                ws.column_dimensions[col_letter].width = max_len + 3

    return buf.getvalue()


def generate_simulation_excel(
    sim_name: str,
    system_size_kw: float,
    dc_ac_ratio: float,
    production_summary: dict | None,
    location_input: str | None,
    lat: float | None,
    lon: float | None,
    system_life_years: int,
    nem_regime_1: str,
    nem_regime_2: str | None,
    num_years_1: int | None,
    battery_capacity_kwh: float,
    discharge_limit_pct: float,
    utility_name: str | None,
    selected_rate_name: str | None,
    rate_escalator_pct: float,
    load_escalator_pct: float,
    annual_projection_df: pd.DataFrame,
    result: BillingResult,
    result_pv_only: BillingResult | None,
    export_rates_8760: "pd.Series | None",
    export_rates_8760_2: "pd.Series | None",
    nem_switch: bool,
    export_rates_multiyear: dict[int, "pd.Series"] | None,
    export_rates_multiyear_2: dict[int, "pd.Series"] | None,
    years: int,
    cod_date=None,
    degradation_pct: float = 0.0,
) -> bytes:
    """Generate a multi-sheet Excel workbook with full simulation details.

    Returns bytes of the .xlsx file content.
    """
    # 1. Summary sheet
    summary_df, summary_rows = _build_summary_df(
        sim_name, system_size_kw, dc_ac_ratio, production_summary,
        location_input, lat, lon, system_life_years,
        nem_regime_1, nem_regime_2, num_years_1,
        battery_capacity_kwh, discharge_limit_pct,
        utility_name, selected_rate_name,
        rate_escalator_pct, load_escalator_pct,
        result, cod_date,
    )

    # 2. Hourly sheets (export rates + retail rates)
    export_hourly_df, retail_hourly_df, _has_bess, exp_kwh, exp_credit, _safe_exp = (
        _build_hourly_sheets(result, battery_capacity_kwh, nem_regime_1, export_rates_8760)
    )

    # 3. Monthly sheets (export rates + retail rates)
    export_monthly_df, retail_monthly_df = _build_monthly_sheets(
        result, result_pv_only, rate_escalator_pct, load_escalator_pct,
        years, export_rates_multiyear, nem_regime_1, nem_regime_2,
        num_years_1, export_rates_multiyear_2, cod_date, degradation_pct,
        _has_bess, exp_kwh, exp_credit, _safe_exp,
    )

    # 4. Annual projection sheet (aggregated from monthly)
    export_annual_df = _build_projection_sheet(export_monthly_df)

    # 5. Assemble and format workbook
    return _write_excel_workbook(
        summary_df, summary_rows, annual_projection_df,
        export_hourly_df, export_monthly_df, export_annual_df,
        retail_hourly_df, retail_monthly_df,
    )
